from datetime import date, datetime

from django.conf import settings
from django.shortcuts import render, redirect
import requests
import json
from django_global_request.middleware import get_request
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from django.contrib import messages
from django.http import HttpResponseRedirect, HttpResponse
from Common.VIEW.Agent.agent_views import upload_visa_doc_get_path, file_upload_get_path, upload_frro_doc_get_path
from Common.models import Corporate_Employee_Login_Access_Token
from landing.cotrav_messeging import Excelexport
from openpyxl import Workbook

def logout_action(request):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']

        access_token = request.session['employee_access_token']
        user = Corporate_Employee_Login_Access_Token.objects.get(access_token=access_token)
        del request.session['employee_login_type']
        del request.session['employee_access_token']

        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        #logout(request)  # the user is now LogOut
        return redirect("/login")
    else:
        return redirect("/login")

def homepage(request):
    if 'employee_login_type' in request.session:
        user_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        payload = {'employee_id': request.user.id, 'corporate_id':request.user.corporate_id}
        url = settings.API_BASE_URL + "employee_dashboard"
        data = getDataFromAPI(user_type, access_token, url, payload)
        dataDashboard = data['Dashboard']

        url_access = settings.API_BASE_URL + "view_company"
        data = getDataFromAPI(user_type, access_token, url_access, payload)
        access = data['Corporates']

        return render(request, 'Company/Employee/home_page.html', {'user': request.user,'dataDashboard':dataDashboard, 'corp_access':access})
    else:
        return HttpResponseRedirect("/login")


def user_profile(request):
    return render(request, 'Company/Employee/user_profile.html', {'user': request.user})


def company_admins(request, id):
    request = get_request()

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            admins = company['Admins']
            return render(request, "Company/Employee/company_admins.html", {'admins': admins})
        else:
            return render(request, "Company/Employee/company_admins.html", {'admins': {}})
    else:
        return HttpResponseRedirect("/login")


def company_billing_entities(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        if company['success'] == 1:
            entities = company['Entitys']
            cities = cities["Cities"]
            return render(request, "Company/Employee/billing_entities.html",{'billing_entities': entities, "cities": cities, })
        else:
            return render(request, "Company/Employee/billing_entities.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_rates(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            company_rates = company['Corporate_Retes']
            return render(request, "Company/Employee/company_rates.html", {'corporate_rates': company_rates})
        else:
            return render(request, "Company/Employee/company_rates.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_groups(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            groups = company['Groups']
            return render(request, "Company/Employee/groups.html", {'groups': groups})
        else:
            return render(request, "Company/Employee/groups.html", {'groups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_subgroups(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            url2 = settings.API_BASE_URL + "groups"
            subgroups = company['Subgroups']
            gr = getDataFromAPI(login_type, access_token, url2, payload)
            groups = gr['Groups']
            return render(request, "Company/Employee/subgroups.html", {'subgroups': subgroups, 'groups': groups})
        else:
            return render(request, "Company/Employee/subgroups.html", {'subgroups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_spocs(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            spocs = company['Spocs']
            return render(request, "Company/Employee/spocs.html", {'spocs': spocs})
        else:
            return render(request, "Company/Employee/spocs.html", {'spocs': {}})
    else:
        return HttpResponseRedirect("/login")


def company_employees(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            employees = company['Employees']
            return render(request, "Company/Employee/employees.html", {'employees': employees})
        else:
            return render(request, "Company/Employee/employees.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/login")


def view_company_group(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_group"
        payload = {'group_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_group_auth"
        grp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(grp_auths)
        if company['success'] == 1:
            groups = company['Groups']
            grp_auths = grp_auths['Groups']
            return render(request, "Company/Employee/view_groups.html", {'group': groups, 'grp_auths': grp_auths})
        else:
            return render(request, "Company/Employee/view_groups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")


def view_company_subgroup(request, id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_subgroup"
        payload = {'subgroup_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        subgrp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(subgrp_auths)
        if company['success'] == 1:
            subgroups = company['SubGroups']
            subgrp_auths = subgrp_auths['SubGroups']
            return render(request, "Company/Employee/view_subgroups.html",
                          {'subgroup': subgroups, 'subgrp_auths': subgrp_auths})
        else:
            return render(request, "Company/Employee/view_subgroups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")

######################################  TAXI  ###################################


def taxi_bookings(request,id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "employee_taxi_bookings"
        payload = {'employee_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/taxi_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Employee/taxi_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_taxi_booking(request,id):

    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/view_taxi_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Employee/view_taxi_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_taxi_booking(request,id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_taxi_booking"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Taxi Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_taxi_booking(request,id):
    if request.method == 'POST':
        if 'employee_login_type' in request.session:
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            current_url = request.POST.get('current_url', '')
            employee_id = request.POST.get('employee_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')

            payload1 = {'spoc_id':spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            print(spoc)

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            tour_type = request.POST.get('tour_type', '')
            pickup_city = request.POST.get('pickup_city', '')
            pickup_location = request.POST.get('pickup_location', '')
            drop_location = request.POST.get('drop_location', '')
            pickup_datetime = request.POST.get('pickup_datetime', '')
            taxi_type = request.POST.get('taxi_type', '')
            package_id = request.POST.get('package_id', '')
            no_of_days = request.POST.get('no_of_days', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            actual_city_id = request.POST.get('current_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = 1

            if tour_type == 1 or tour_type == '1':
                url_add_city = settings.API_BASE_URL + "add_city_name"
                pickup_details = [x.strip() for x in pickup_city.split(',')]
                city_data = {'login_type': login_type, 'access_token': access_token, 'city_name': pickup_details[0], 'state_id': '1'}
                city_id = getDataFromAPI(login_type, access_token, url_add_city, city_data)
                for conty_id in city_id['id']:
                    actual_city_id = conty_id['id']

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id', ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':employee_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'tour_type':tour_type,'pickup_city':actual_city_id,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'pickup_location':pickup_location,'drop_location':drop_location,'pickup_datetime':pickup_datetime+':00','taxi_type':taxi_type,
                       'package_id':package_id,'no_of_days':no_of_days,'reason_booking':reason_booking,'no_of_seats':no_of_seats,
                       'employees':employees,'entity_id':entity_id,'is_sms':1,'is_email':1}

            url_taxi_booking = settings.API_BASE_URL + "add_taxi_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Employee/taxi-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed To Add Taxi Booking..!')
                return HttpResponseRedirect("/Corporate/Employee/taxi-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'employee_login_type' in request.session:
            request = get_request()
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_city1 = settings.API_BASE_URL + "city_by_package"
            cities1 = getDataFromAPI(login_type, access_token, url_city1, payload)
            citiess = cities1['Cities']

            url_taxi = settings.API_BASE_URL + "taxi_types"
            taxies = getDataFromAPI(login_type, access_token, url_taxi, payload)
            taxies = taxies['taxi_types']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Employee/add_taxi_booking.html', {'employees':employees,'cities':cities,'taxies':taxies,
                    'assessments':ass_code,'citiess':citiess, 'corp_access':access})
            else:
                return render(request, 'Company/Employee/add_taxi_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


############################################## BUS ##########################################


def bus_bookings(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "employee_bus_bookings"
        payload = {'employee_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/bus_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Employee/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_bus_booking(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Employee/view_bus_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_bus_booking(request,id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_bus_booking"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Bus Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_bus_booking(request,id):
    if request.method == 'POST':
        if 'employee_login_type' in request.session:
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            current_url = request.POST.get('current_url', '')
            employee_id = request.POST.get('employee_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')

            payload1 = {'spoc_id':spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            print(spoc)

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            bus_type = request.POST.get('bus_type', '')
            bus_type2 = request.POST.get('bus_type2', '')
            bus_type3 = request.POST.get('bus_type3', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            preferred_board_point = request.POST.get('preferred_board_point', '')
            preferred_drop_point = request.POST.get('preferred_drop_point', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = 1

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id', ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': employee_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from': from_location, 'to': to_location,
                       'bus_type': bus_type, 'booking_datetime': booking_datetime, 'journey_datetime': journey_datetime+':00','journey_datetime_to':journey_datetime_to+':00',
                       'entity_id': entity_id,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'preferred_bus': preferred_bus, 'reason_booking': reason_booking, 'no_of_seats': no_of_seats,
                       'employees': employees,'is_sms':1,'is_email':1,'preferred_board_point':preferred_board_point,
                       'preferred_drop_point':preferred_drop_point,'bus_type2':bus_type2,'bus_type3':bus_type3}

            url_taxi_booking = settings.API_BASE_URL + "add_bus_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Employee/bus-bookings/2", {'message': "Operation Successfully"})
            else:
                return HttpResponseRedirect("/Corporate/Employee/bus-bookings/2", {'message': "Operation Failed"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'employee_login_type' in request.session:
            request = get_request()
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']

            payload = {'corporate_id': id}

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Employee/add_bus_booking.html', {'cities':cities,'assessments':ass_code, 'corp_access':access})
            else:
                return render(request, 'Company/Employee/add_bus_booking.html', {})
        else:
            return HttpResponseRedirect("/login")



############################################## Train ##########################################


def train_bookings(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "employee_train_bookings"
        payload = {'employee_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/train_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Employee/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_train_booking(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Employee/view_train_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_train_booking(request,id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_train_booking"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Train Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')
    

def add_train_booking(request,id):
    if request.method == 'POST':
        if 'employee_login_type' in request.session:
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            current_url = request.POST.get('current_url', '')
            employee_id = request.POST.get('employee_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')

            payload1 = {'spoc_id':spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            train_type = request.POST.get('train_type', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = 1

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id', ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': employee_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from': from_location, 'to': to_location,'journey_datetime_to':journey_datetime_to+':00',
                       'train_type': train_type, 'booking_datetime': booking_datetime, 'journey_datetime': journey_datetime+':00',
                       'entity_id': entity_id,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'preferred_bus': preferred_bus, 'reason_booking': reason_booking, 'no_of_seats': no_of_seats,
                       'employees': employees,'is_sms':1,'is_email':1}

            url_taxi_booking = settings.API_BASE_URL + "add_train_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Employee/train-bookings/2", {})
            else:
                messages.error(request,'Failed To Add Train Booking')
                return HttpResponseRedirect("/Corporate/Employee/train-bookings/2", {})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'employee_login_type' in request.session:
            request = get_request()
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            payload = {'corporate_id': id}

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_train = settings.API_BASE_URL + "train_types"
            trains = getDataFromAPI(login_type, access_token, url_train, payload)
            types = trains['Types']
            print(types)

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            # url_railway_stations = settings.API_BASE_URL + "railway_stations"
            # trains1 = getDataFromAPI(login_type, access_token, url_railway_stations, payload)
            # railway_stations = trains1['Stations']
            railway_stations = ""

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Employee/add_train_booking.html', {'cities':cities,'types':types,'assessments':ass_code,
                                                                                   'railway_stations':railway_stations, 'corp_access':access})
            else:
                return render(request, 'Company/Employee/add_train_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


############################################## HOTELS ##########################################


def hotel_bookings(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "employee_hotel_bookings"
        payload = {'employee_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/hotel_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Employee/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_hotel_booking(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Employee/view_hotel_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_hotel_booking(request,id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_hotel_booking"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Hotel to Reject Train Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_hotel_booking(request,id):
    if request.method == 'POST':
        if 'employee_login_type' in request.session:
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('employee_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')

            payload1 = {'spoc_id':spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            from_city = request.POST.get('from_city')
            city_area = request.POST.get('city_area')
            preferred_hotel_area = request.POST.get('preferred_hotel_area')
            check_in_date = request.POST.get('check_in_date')
            check_out_date = request.POST.get('check_out_date')
            room_type_priority1 = request.POST.get('room_type_priority1')
            room_type_priority2 = request.POST.get('room_type_priority2')
            room_occupancy = request.POST.get('room_occupancy')
            preferred_hotel = request.POST.get('preferred_hotel')
            booking_date = request.POST.get('booking_datetime')
            no_of_nights = request.POST.get('no_of_nights', '')
            assessment_code = request.POST.get('assessment_code')

            assessment_city = request.POST.get('assessment_city')
            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')

            if billing_entity:
                pass
            else:
                billing_entity=0

            employees = []

            for i in range(1,2):
                employees.append(request.POST.get('employee_id', ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from_city_id': from_city, 'from_area_id': city_area,
                       'preferred_area': preferred_hotel_area, 'checkin_datetime': check_in_date+':00',
                       'checkout_datetime': check_out_date+':00', 'bucket_priority_1': room_type_priority1,
                       'bucket_priority_2': room_type_priority2, 'room_type_id': room_occupancy,
                       'preferred_hotel': preferred_hotel, 'booking_datetime': booking_date,
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city,
                       'billing_entity_id': billing_entity, 'employees': employees,'no_of_nights':no_of_nights,
                       'reason_booking':reason_for_booking,'no_of_seats':1,'is_sms':1,'is_email':1}

            url_taxi_booking = settings.API_BASE_URL + "add_hotel_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Employee/hotel-bookings/2", {})
            else:
                messages.error(request, 'Failed To Add Hotel Booking..!')
                return HttpResponseRedirect("/Corporate/Employee/hotel-bookings/2", {})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'employee_login_type' in request.session:
            request = get_request()
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']

            payload = {'corporate_id': id}
            url_room_types = settings.API_BASE_URL + "room_types"
            room_types = getDataFromAPI(login_type, access_token, url_room_types, payload)
            room_types = room_types['Types']

            url_hotel_types = settings.API_BASE_URL + "hotel_types"
            hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
            hotel_types = hotel_types['Types']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities_ass = cities['AssCity']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            # url_city1 = settings.API_BASE_URL + "cities"
            # cities1 = getDataFromAPI(login_type, access_token, url_city1, payload)
            # cities = cities1['Cities']
            cities = ""

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Employee/add_hotel_booking.html', {'room_types':room_types,'cities_ass':cities_ass,
                                    'hotel_types':hotel_types,'assessments':ass_code,'cities':cities, 'corp_access':access})
            else:
                return render(request, 'Company/Employee/add_hotel_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

############################################## FLIGHT ##########################################


def flight_bookings(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "employee_flight_bookings"
        payload = {'employee_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/flight_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Employee/flight_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_flight_booking(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/view_flight_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Employee/view_flight_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_flight_booking(request,id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_flight_booking"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Flight Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_flight_booking(request,id):
    if request.method == 'POST':
        if 'employee_login_type' in request.session:
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('employee_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')

            payload1 = {'spoc_id':spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']
            group_id =0
            subgroup_id =0

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            usage_type = request.POST.get('usage_type', '')
            trip_type = request.POST.get('trip_type', '')
            seat_type = request.POST.get('seat_type', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            departure_date = request.POST.get('departure_date', '')
            preferred_flight = request.POST.get('preferred_flight', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = 1

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id', ''))
                print(employees)

            payload = {'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'usage_type':usage_type,'trip_type':trip_type,'seat_type':seat_type,'from_city':from_city,'to_city':to_city,
                       'booking_datetime':booking_datetime,'departure_datetime':departure_date,'preferred_flight':preferred_flight,'assessment_code':assessment_code,
                       'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'billing_entity_id':entity_id,
                       'assessment_city_id':assessment_city_id,'is_sms':1,'is_email':1}

            url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Employee/flight-bookings/2", {})
            else:
                messages.error(request,'Failed To Add Flight Booking')
                return HttpResponseRedirect("/Corporate/Employee/flight-bookings/2", {})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'employee_login_type' in request.session:
            request = get_request()
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            payload = {'corporate_id': id}

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities_ass = cities['AssCity']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            if id:
                return render(request, 'Company/Employee/add_flight_booking.html', {'cities_ass':cities_ass,'assessments':ass_code, 'corp_access':access,
                'airports':airports})
            else:
                return render(request, 'Company/Employee/add_flight_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


####################### Download MIS ##################################

def dateonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError:
        return ''


def download_taxi_bookings(request):
    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

        # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [

        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',

        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'Passengers',

        'Booking Date',
        'Booking Time',

        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        'Approver1 Action',

        'Approver1 Name',

        'Approver1 Date',

        'Approver1 Time',

        'Approver2 Action',

        'Approver2 Name',

        'Approver2 Date',

        'Approver2 Time',

        'Approved Date',

        'Approved Time',

        'Approved By',

        'Canceled Date',

        'Canceled Time',

        'Canceled By',

        'Assigned Date',

        'Assigned Time',

        'Assigned By',

        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',

        'Package Name',
        'Tour Type',
        'Vehicle Type',

        'Driver Name',
        'Driver Contact	',
        'Taxi Reg No.',
        'No. Of Seats',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        spoc_status = ''
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by = ''
        spoc_canceled_date = ''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

            if bk['spoc_status'] == 1:
                spoc_status = "In-Active"
            else:
                spoc_status = "Active"

            if (act['action'] == 3 and act['user_type'] == 6):
                print('canceled')
                spoc_canceled_by = act['employee_name']
                spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [
            bk['reference_no'],
            bk['assessment_city_id'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],

            bk['zone_name'],
            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),
            dateonly(bk['pickup_datetime']),
            timeonly(bk['pickup_datetime']),

            bk['package_name'],
            bk['tour_type'],
            bk['taxi_type_name'],

            bk['driver_name'],
            bk['driver_contact'],
            '',
            bk['no_of_seats'],

            bk['client_status'],

            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Bus Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Reason For Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",
        'Approver1 Action',
        'Approver1 Name',
        'Approver1 Date',
        'Approver1 Time',
        'Approver2 Action',
        'Approver2 Name',
        'Approver2 Date',
        'Approver2 Time',
        'Approved Date',
        'Approved Time',
        'Approved By',
        'Canceled Date',
        'Canceled Time',
        'Canceled By',
        'Assigned Date',
        'Assigned Time',
        'Assigned By',
        'Passanger Name',
        'Pickup City',
        'Drop City',
        'Journey Date',
        'Journey Time',
        'Current Booking Status',
        'Bus Type Allocated',
        'PNR Number',
        'Ticket Price',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by = ''
        spoc_canceled_date = ''
        spoc_status = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

                # Define the data for each cell in the row
        row = [

            bk['reference_no'],

            bk['assessment_code'],
            bk['assessment_city_id'],

            bk['reason_booking'],

            bk['zone_name'],

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            dateonly(bk['booking_datetime']),

            timeonly(bk['booking_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            passanger_list,

            bk['pickup_location'],

            bk['drop_location'],

            dateonly(bk['pickup_from_datetime']),

            timeonly(bk['pickup_from_datetime']),

            bk['cotrav_status'],

            bk['assign_bus_type_id'],

            bk['pnr_no'],

            bk['client_status'],
            bk['cotrav_status'],

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "Pickup City",
        "Drop City",
        "Booking Date",
        "Booking Time",
        "Journey Date",
        "Journey Time",
        "SPOC Status",
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        "Zone",
        "Coach Type Allocated",
        "Quota Used",

        'No of seats',
        'Operator name',
        'Operator contact',
        'Train name',
        'Ticket no',
        'pnr no',
        'Assign bus type id',
        'Seat no',
        'Portal used',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        spoc_canceled_by = ''
        spoc_canceled_date = ''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            bk['zone_name'],
            bk['train_type_priority_1'],
            bk['seat_no'],

            bk['no_of_seats'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['train_name'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['seat_no'],
            bk['portal_used'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "From City",
        "To City",
        "Booking Date",
        "Booking Time",
        "Departure Date",
        "Departure Time",

        "Return Date",
        "Booking Status",

        "SPOC Status",
        "SPOC Cancel By",
        "SPOC Cancel Date",

        "Usage Type",
        "Trip Type",
        "Flight Type",
        "Seat Type",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        'First Flight Name',
        'First Flight No',
        'First Flight PNR No.',
        'First Flight From',
        'First Flight To',
        'First Flight Departure Date',
        'First Flight Departure time',
        'First Flight Arrival Date',
        'First Flight Arrival time',

        'Second Flight Name',
        'Second Flight No',
        'Second Flight PNR No.',
        'Second Flight From',
        'Second Flight To',
        'Second Flight Departure Date',
        'Second Flight Departure Time',
        'Second Flight Arrival Datetime',
        'Second Flight Arrival Time',

        'Third Flight Name',
        'Third Flight No',
        'Third Flight PNR No.',
        'Third Flight From',
        'Third Flight To',
        'Third Flight Departure Date',
        'Third Flight Departure Time',
        'Third Flight Arrival Date',
        'Third Flight Arrival Time',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        spoc_status = ''
        spoc_canceled_by = ''
        spoc_canceled_date = ''

        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            "Booking Remarks",
            bk['from_location'],
            bk['to_location'],

            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),

            "",
            bk['cotrav_status'],

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            flight_name1,
            flight_no1,
            pnr_no1,
            from_city1,
            to_city1,
            dateonly(departure_datetime1),
            timeonly(departure_datetime1),
            dateonly(arrival_datetime1),
            timeonly(arrival_datetime1),

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            dateonly(departure_datetime2),
            timeonly(departure_datetime2),
            dateonly(arrival_datetime2),
            timeonly(arrival_datetime2),

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            dateonly(departure_datetime3),
            timeonly(departure_datetime3),
            dateonly(arrival_datetime3),
            timeonly(arrival_datetime3),

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Bucket Request Date',
        'Bucket Request Time',
        'Bucket Approved Date',
        'Bucket Approved Time',
        'Preferred Hotel',
        'Assigned Hotel',
        'Assigned Hotel Address',
        'Hotel Contact',
        'Assign Date',
        'Assign Time',
        'TaxiVaxi Status',
        'Employees Name',
        'No. of Persons',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Higher Bucket Requested',
        'Reason for Higher Bucket',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',

        'Is Prepaid',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"
        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            'Billing Entity',
            'Travel request Code',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            '',
            '',
            '',
            '',

            bk['preferred_hotel'],
            bk['assign_hotel_id'],
            bk['operator_name'],
            bk['operator_contact'],
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['status_cotrav'],
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),

            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def visa_bokings(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        if request.method == 'POST':
            pass
        else:
            url = settings.API_BASE_URL+"employee_visa_services"
            payload = {'user_id': request.user.id}
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                corporates_data = company['Visa']
                return render(request,"Company/Employee/visa_requests.html",{'visa_services':corporates_data})
            else:
                return render(request,"Company/Employee/visa_requests.html",{'visa_services':{}})
    else:
        return HttpResponseRedirect("/login")


def view_visa_requests(request, id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        if request.method == 'POST':
            current_url = request.POST.get('current_url')
            request_type = request.POST.get('request_type')
            visa_request_type = request.POST.get('visa_request_type')
            corporate_id = request.POST.get('corporate_id')
            spoc_id = request.POST.get('spoc_id')
            group_id = request.POST.get('group_id')
            subgroup_id = request.POST.get('subgroup_id')
            country_id = request.POST.get('country_id')
            visa_type = request.POST.get('visa_type')
            purpose_of_trip = request.POST.get('purpose_of_trip')
            current_country_id = request.POST.get('current_country_id')
            current_state_id = request.POST.get('current_state_id')
            consulate_office_id = request.POST.get('consulate_office_id')
            no_of_employees = request.POST.get('no_of_employees')
            employee_email_1 = request.POST.get('employee_email_1')
            employee_name_1 = request.POST.get('employee_name_1')

            payload = {'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'country_id': country_id, 'visa_type': visa_type,
                       'purpose_of_trip': purpose_of_trip, 'current_country_id': current_country_id,
                       'current_state_id': current_state_id, 'consulate_office_id': consulate_office_id,
                       'no_of_employees': no_of_employees, 'employee_name_1': employee_name_1,
                       'employee_email_1': employee_email_1,'visa_request_type':visa_request_type,'request_type':request_type}

            url_cities1 = settings.API_BASE_URL + "add_visa_requests"
            company = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Visa Service Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Visa Service ..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Fails"})

        else:
            payload = {'some': 'data', 'spoc_id': request.user.id, 'visa_id':id}

            url_cities1 = settings.API_BASE_URL + "view_visa_request"
            taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
            visas = taxies11sad['Visa']

            url_cities1 = settings.API_BASE_URL + "get_countries"
            taxies11 = getDataFromAPI(login_type, access_token, url_cities1, payload)
            Country = taxies11['Country']

            url_cities111 = settings.API_BASE_URL + "get_states"
            taxies1ds1 = getDataFromAPI(login_type, access_token, url_cities111, payload)
            states = taxies1ds1['State']

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']

            url_emp = settings.API_BASE_URL + "spoc_employee"
            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employees']

            return render(request, "Company/Employee/view_visa_request.html", {'visas':visas,'countrys': Country, 'states': states, 'cities': cities, 'employees': employees})

    else:
        return HttpResponseRedirect("/login")


def add_visa_requests(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        if request.method == 'POST':
            current_url = request.POST.get('current_url')
            request_type = request.POST.get('request_type')
            visa_request_type = request.POST.get('visa_request_type')
            corporate_id = request.POST.get('corporate_id')
            billing_entity_id = request.POST.get('billing_entity_id')
            spoc_id = request.POST.get('spoc_id')
            group_id = 0
            subgroup_id = 0

            payload1 = {'spoc_id': spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']


            country_id = request.POST.get('country_id')
            visa_type = request.POST.get('visa_type')
            visa_duration = request.POST.get('visa_duration')
            purpose_of_trip = request.POST.get('purpose_of_trip')
            if purpose_of_trip == 'Other':
                purpose_of_trip = request.POST.get('purpose_of_trip_txt')
            current_country_id = request.POST.get('current_country_id')
            current_state_id = request.POST.get('current_state_id')
            consulate_office_id = request.POST.get('consulate_office_id')
            no_of_employees = request.POST.get('no_of_employees')
            employee_email_1 = request.POST.get('employee_email_1')
            no_of_family_member = request.POST.get('no_of_family_member')
            application_form_link = request.POST.get('application_form_link')

            employee_name_1 = request.POST.getlist('employee_id')

            emp_no_of_document = request.POST.getlist('emp_no_of_document')
            emp_document = request.POST.getlist('emp_document')
            emp_document_txt = request.POST.getlist('emp_document_txt')

            family_members_name = request.POST.getlist('family_members_name')
            family_members_relationship = request.POST.getlist('family_members_relationship')
            empf_no_of_document = request.POST.getlist('empf_no_of_document')
            empf_document = request.POST.getlist('empf_document')
            empf_document_path = request.POST.getlist('empf_document_path')
            empf_document_txt = request.POST.getlist('empf_document_txt')

            employee_application_form = []
            employee_docs = []
            employeef_application_form = []
            employeef_docs = []

            final_emp_no = int(no_of_employees) + 1
            final_no_of_family = int(no_of_family_member) + 1
            if request.FILES:
                if request_type == 'Family':
                    for i in range(1, final_no_of_family):
                        if request.FILES['empf_application_' + str(i)] != '':
                            file_up2 = request.FILES['empf_application_' + str(i)]
                            booking_email1 = upload_visa_doc_get_path(file_up2)
                            employeef_application_form.append(booking_email1)
                            val_i = i - 1
                            finl_emp_no_of_document = int(empf_no_of_document[val_i]) + 1
                            for ii in range(1, finl_emp_no_of_document):
                                if request.FILES['empf_document_path_' + str(i) + "_" + str(ii)] != '':
                                    file_up11 = request.FILES['empf_document_path_' + str(i) + "_" + str(ii)]
                                    employee_docs11 = upload_visa_doc_get_path(file_up11)
                                    employeef_docs.append(employee_docs11)
                else:
                    for i in range(1, final_emp_no):
                        if request.FILES['emp_application_form_' + str(i)] != '':
                            file_up = request.FILES['emp_application_form_' + str(i)]
                            booking_email1 = upload_visa_doc_get_path(file_up)
                            employee_application_form.append(booking_email1)
                            val_i = i - 1
                            finl_emp_no_of_document = int(emp_no_of_document[val_i]) + 1

                            for ii in range(1, finl_emp_no_of_document):
                                if request.FILES['emp_document_path_' + str(i) + "_" + str(ii)] != '':
                                    file_up11 = request.FILES['emp_document_path_' + str(i) + "_" + str(ii)]
                                    employee_docs11 = upload_visa_doc_get_path(file_up11)
                                    employee_docs.append(employee_docs11)

            payload = {'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'country_id': country_id, 'visa_type': visa_type,
                       'purpose_of_trip': purpose_of_trip, 'current_country_id': current_country_id,
                       'current_state_id': current_state_id, 'consulate_office_id': consulate_office_id,
                       'no_of_employees': no_of_employees, 'employee_ids': employee_name_1,
                       'employee_email_1': employee_email_1, 'request_type': request_type,
                       'visa_request_type': visa_request_type, 'visa_duration': visa_duration,
                       'no_of_family_member': no_of_family_member, 'application_form_link': application_form_link,
                       'emp_no_of_document': emp_no_of_document, 'emp_document': emp_document,
                       'emp_document_txt': emp_document_txt, 'family_members_name': family_members_name,
                       'family_members_relationship': family_members_relationship,
                       'empf_no_of_document': empf_no_of_document, 'empf_document': empf_document,
                       'empf_document_path': empf_document_path,
                       'employee_application_form': employee_application_form, 'employee_docs': employee_docs,
                       'employeef_application_form': employeef_application_form, 'employeef_docs': employeef_docs,
                       'empf_document_txt': empf_document_txt, 'is_email': 1, 'is_sms': 1, 'billing_entity_id':billing_entity_id}
            print(payload)
            url_cities1 = settings.API_BASE_URL + "add_visa_requests"
            company = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Visa Service Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Visa Service ..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Fails"})

        else:
            payload = {'some': 'data', 'spoc_id': request.user.id}

            url_cities1 = settings.API_BASE_URL + "get_country_provided_by_cotrav"
            taxies11 = getDataFromAPI(login_type, access_token, url_cities1, payload)
            Country = {}
            if taxies11['success'] == 1:
                Country = taxies11['Country']
            else:
                Country = {}

            return render(request, "Company/Employee/add_visa_request.html", {'countrys': Country})

    else:
        return HttpResponseRedirect("/login")


def add_booking_feedback(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_id = request.POST.get('user_id', '')
        booking_type = request.POST.get('booking_type', '')
        user_rating = request.POST.get('user_rating', '')

        user_actual_rating = ''
        if int(user_rating) >= 4:
            user_rating_text_good = request.POST.get('user_rating_text_good', '')
            if user_rating_text_good == 'other':
                user_actual_rating = request.POST.get('user_rating_text_good_txt', '')
            else:
                user_actual_rating = user_rating_text_good
        else:
            user_rating_text_bad = request.POST.get('user_rating_text_bad', '')
            if user_rating_text_bad == 'other':
                user_actual_rating = request.POST.get('user_rating_text_bad_txt', '')
            else:
                user_actual_rating = user_rating_text_bad

        url = settings.API_BASE_URL + "add_booking_feedback"
        payload = {'booking_id': booking_id, 'user_id': user_id, 'user_actual_rating': user_actual_rating, 'booking_type':booking_type, 'user_rating':user_rating}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            messages.error(request, company['message'])
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, company['message'])
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def view_frro_request(request, id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        if request.method == 'POST':
            pass
        else:
            payload = {'frro_id': id, 'user_id': request.user.id}
            url_cities1 = settings.API_BASE_URL + "view_frro_request"
            taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(taxies11sad)
            if taxies11sad['success'] == 1:
                visas = taxies11sad['Frro']
                return render(request,"Company/Employee/view_frro_request.html",{'visas':visas})
            else:
                return render(request,"Company/Employee/view_frro_request.html",{'visas':{}})
    else:
        return HttpResponseRedirect("/login")


def get_all_frro_requests(request, id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        if request.method == 'POST':
            pass
        else:
            url = settings.API_BASE_URL + "company_frro_request"
            payload = {'request_type': id, 'user_id': request.user.id}
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                corporates_data = company['Frros']
                return render(request, "Company/Employee/frro_services.html", {'visa_services': corporates_data, 'booking_type': id})
            else:
                return render(request, "Company/Employee/frro_services.html", {'visa_services': {}})
    else:
        return HttpResponseRedirect("/login")


def add_new_frro_request(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        if request.method == 'POST':
            current_url = request.POST.get('current_url')
            request_type = request.POST.get('country_id')
            corporate_id = request.POST.get('corporate_id')
            billing_entity_id = request.POST.get('billing_entity_id')

            group_id = 0
            subgroup_id = 0
            spoc_id = request.POST.get('spoc_id')
            payload1 = {'spoc_id': spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            employee_id = request.POST.get('employee_id')

            city_id = request.POST.get('city_id')
            registered_email = request.POST.get('registered_email')
            reason_of_booking = request.POST.get('reason_of_booking')
            registered_username = request.POST.get('registered_username')
            registered_password = request.POST.get('registered_password')

            father_name = request.POST.get('father_name')
            mother_name = request.POST.get('mother_name')
            spouse_name = request.POST.get('spouse_name')
            height = request.POST.get('height')
            relegion = request.POST.get('relegion')
            outind_address = request.POST.get('outind_address')
            outind_city = request.POST.get('outind_city')
            outind_country = request.POST.get('outind_country')
            ind_address = request.POST.get('ind_address')
            ind_city = request.POST.get('ind_city')
            ind_country = request.POST.get('ind_country')
            embarkation_place = request.POST.get('embarkation_place')
            embarkation_city = request.POST.get('embarkation_city')
            embarkation_country = request.POST.get('embarkation_country')
            date_of_arrival = request.POST.get('date_of_arrival')
            disembarkation_place = request.POST.get('disembarkation_place')
            mode_of_journey = request.POST.get('mode_of_journey')
            trip_number = request.POST.get('trip_number')
            purpose_of_visit = request.POST.get('purpose_of_visit')
            is_military = request.POST.get('is_military')

            passport_no = request.POST.get('passport_no')
            passport_issue_date = request.POST.get('passport_issue_date')
            passport_expiry_date = request.POST.get('passport_expiry_date')
            visa_number = request.POST.get('visa_number')
            visa_expiry_date = request.POST.get('visa_expiry_date')
            current_location = request.POST.get('current_location')
            current_stay_address = request.POST.get('current_stay_address')
            company_contract_expiry_date = request.POST.get('company_contract_expiry_date')
            formc_expiry_date = request.POST.get('formc_expiry_date')
            lease_status = request.POST.get('lease_status')

            no_of_employees = 1
            is_email = request.POST.get('is_email')
            is_sms = request.POST.get('is_sms')

            employee_docs = {}
            doc_files = {}
            qty_docs = request.POST.getlist('doc_select[]')
            file_up11 = request.FILES.getlist('document[]')
            final_emp_no = int(no_of_employees) + 1
            # print(request.FILES['document[]'])
            # print(file_up11)
            for i, val in enumerate(qty_docs):
                if (0 <= i) and (i < len(file_up11)):
                    print(i)
                    print(file_up11[i])
                    employee_docs11 = upload_frro_doc_get_path(file_up11[i])
                    key = qty_docs[i]
                    employee_docs[key] = employee_docs11
                    doc_files[key] = employee_docs11
                else:
                    key = qty_docs[i]
                    doc_files[key] = ""

            payload = {'user_id': request.user.id, 'user_type': login_type, 'corporate_id': corporate_id,
                       'spoc_id': spoc_id,
                       'is_email': is_email, 'is_sms': is_sms, 'group_id': group_id, 'entity_id': billing_entity_id,
                       'no_of_employees': no_of_employees, 'subgroup_id': subgroup_id, 'request_type': request_type,
                       'employee_ids': employee_id, 'employee_docs': json.dumps(employee_docs),
                       'doc_files': json.dumps(doc_files), 'father_name': father_name, 'mother_name': mother_name,
                       'spouse_name': spouse_name, 'height': height, 'relegion': relegion,
                       'outind_address': outind_address, 'outind_city': outind_city, 'outind_country': outind_country,
                       'ind_address': ind_address, 'ind_city': ind_city, 'ind_country': ind_country,
                       'embarkation_place': embarkation_place, 'embarkation_city': embarkation_city,
                       'embarkation_country': embarkation_country, 'date_of_arrival': date_of_arrival,
                       'disembarkation_place': disembarkation_place, 'mode_of_journey': mode_of_journey,
                       'trip_number': trip_number, 'purpose_of_visit': purpose_of_visit, 'is_military': is_military,
                       'city_id': city_id, 'registered_email': registered_email,
                       'reason_of_booking': reason_of_booking, 'registered_username': registered_username,
                       'registered_password': registered_password ,'passport_no':passport_no,
                       'passport_issue_date':passport_issue_date,'passport_expiry_date':passport_expiry_date,'visa_number':visa_number,'visa_expiry_date':visa_expiry_date,
                       'current_location':current_location,'current_stay_address':current_stay_address,'company_contract_expiry_date':company_contract_expiry_date,
                       'formc_expiry_date':formc_expiry_date,'lease_status':lease_status}
            print(payload)
            url_cities1 = settings.API_BASE_URL + "add_frro_requests"
            company = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, company['message'])
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect(current_url, {'message': "Operation Fails"})

        else:
            payload = {'some': 'data', 'spoc_id': request.user.id, 'employee_id': request.user.id}

            url_cities1 = settings.API_BASE_URL + "get_frro_request_type"
            taxies11 = getDataFromAPI(login_type, access_token, url_cities1, payload)
            req_types = {}

            if taxies11['success'] == 1:
                req_types = taxies11['Types']
            else:
                req_types = {}

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities["Cities"]

            url_city1 = settings.API_BASE_URL + "get_employee_questionnaire_details"
            emp = getDataFromAPI(login_type, access_token, url_city1, payload)
            emp_ques = emp["Details"]


            return render(request, "Company/Employee/add_frro_request.html", {'req_types': req_types, 'cities': cities, 'emp_ques':emp_ques})

    else:
        return HttpResponseRedirect("/login")


def frro_change_document(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        current_url = request.POST.get('current_url', '')
        document_id = request.POST.get('document_id', '')

        global booking_email
        booking_email = ''
        if request.FILES:
            file_up = request.FILES.get('documnet_file', False)
            if file_up:
                file_up = request.FILES['documnet_file']
                booking_email = upload_frro_doc_get_path(file_up)
            else:
                booking_email = None
        else:
            booking_email = None

        payload = {'document_id': document_id, 'document_file': booking_email}
        print(payload)
        url_cities1 = settings.API_BASE_URL + "frro_change_document"
        taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
        return HttpResponseRedirect(current_url, {})

    else:
        return HttpResponseRedirect("/login")


def frro_change_document_letter(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        booking_id_main = request.POST.get('booking_id_main', '')

        global booking_email
        booking_email = ''
        if request.FILES:
            file_up = request.FILES.get('booking_email', False)
            if file_up:
                file_up = request.FILES['booking_email']
                booking_email = upload_frro_doc_get_path(file_up)
            else:
                booking_email = None
        else:
            booking_email = None

        payload = {'booking_id': booking_id, 'booking_id_main':booking_id_main, 'document_file': booking_email}
        print(payload)
        print(current_url)
        url_cities1 = settings.API_BASE_URL + "frro_change_document_letter"
        taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
        print(taxies11sad)
        return HttpResponseRedirect(current_url, {})

    else:
        return HttpResponseRedirect("/login")


def frro_change_office_status(request):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        booking_id_main = request.POST.get('booking_id_main', '')
        status_type = request.POST.get('status_type', '')
        status = request.POST.get('status', '')
        status_comment = request.POST.get('status_comment', '')

        global booking_email
        booking_email = ''
        if request.FILES:
            file_up = request.FILES.get('booking_email', False)
            if file_up:
                file_up = request.FILES['booking_email']
                booking_email = upload_frro_doc_get_path(file_up)
            else:
                booking_email = None
        else:
            booking_email = None

        payload = {'booking_id': booking_id, 'booking_id_main':booking_id_main, 'document_file': booking_email, 'status_type':status_type, 'status':status, 'status_comment':status_comment}
        print(payload)
        print(current_url)
        url_cities1 = settings.API_BASE_URL + "frro_change_office_status"
        taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
        print(taxies11sad)
        return HttpResponseRedirect(current_url, {})

    else:
        return HttpResponseRedirect("/login")


def view_employee_frro_details(request, id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        payload = {'employee_id': id}
        url = settings.API_BASE_URL + "view_employee"
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = company['Employee']

        url = settings.API_BASE_URL + "view_employee"
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = company['Employee']

        url2 = settings.API_BASE_URL + "view_employee_documents"
        company1 = getDataFromAPI(login_type, access_token, url2, payload)
        companies1 = company1['Employee']

        return render(request, "Company/Spoc/view_employee_document_details.html", {'employees': companies, 'documents':companies1})

    else:
            return HttpResponseRedirect("/login")


############################################## GUEST HOUSE ##########################################


def guesthouse_bookings(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "company_guesthouse_booking"
        payload = {'user_id': user_id, 'request_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/guesthouse_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Employee/guesthouse_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def search_guesthouse(request, id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        payload = {'user_id':user_id, 'corporate_id':request.user.corporate_id}

        url_hotel_types = settings.API_BASE_URL + "get_guesthouse_by_corporate"
        hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)

        if hotel_types['success'] == 1:
            guesthouse = hotel_types['Guesthouse']
            return render(request, "Company/Employee/search_guesthouses.html",{'guesthouses': guesthouse})
        else:
            return render(request, "Company/Employee/search_guesthouses.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_guesthouse_booking(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        url = settings.API_BASE_URL + "view_guesthouse_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Employee/view_guesthouse_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Employee/view_guesthouse_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_guesthouse_booking(request,id):
    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_guesthouse_booking"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Guesthouse Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Guesthouse to Reject Train Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def get_guesthouse_to_booking(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        gid = request.POST.get('guesthouse_id', '')
        from_city = request.POST.get('from_city', '')
        check_in = request.POST.get('check_in_datetime', '')
        check_out = request.POST.get('check_out_datetime', '')
        gid = request.POST.get('guesthouse_id', '')
        rid = request.POST.get('room_id', '')
        no_of_nights = request.POST.get('no_of_nights', '')

        payload = {'corporate_id': request.user.corporate_id, 'guesthouse_id': gid}

        url_hotel_types = settings.API_BASE_URL + "get_guesthouse_details"
        hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
        guesthouse = hotel_types['Guesthouse']

        url_city = settings.API_BASE_URL + "get_assessment_city"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        cities_ass = cities['AssCity']

        url_ass_code = settings.API_BASE_URL + "get_assessment_code"
        ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
        ass_code = ass_code['AssCodes']

        url_access = settings.API_BASE_URL + "view_company"
        data = getDataFromAPI(login_type, access_token, url_access, payload)
        access = data['Corporates']

        if gid:
            return render(request, 'Company/Employee/add_guesthouse_booking.html', {'cities_ass':cities_ass,'guesthouses':guesthouse,'no_of_nights':no_of_nights,
           'assessments':ass_code,'cities':cities, 'corp_access':access,'from_city':from_city,'check_in':check_in,'check_out':check_out,'gid':gid,'rid':rid})
        else:
            return render(request, 'Company/Employee/add_guesthouse_booking.html', {})
    else:
        return HttpResponseRedirect("/login")


def add_guesthouse_booking(request,id):
    if request.method == 'POST':
        if 'employee_login_type' in request.session:
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('employee_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id =0
            subgroup_id=0
            payload1 = {'spoc_id':spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            room_id = request.POST.get('room_id')
            guesthouse_id = request.POST.get('guesthouse_id')

            city_id = request.POST.get('city_id')
            booking_date = request.POST.get('booking_datetime')
            checkin_datetime = request.POST.get('checkin_datetime', '')
            checkout_datetime = request.POST.get('checkout_datetime', '')
            no_of_nights = request.POST.get('no_of_nights', '')
            assessment_code = request.POST.get('assessment_code')
            assessment_city = request.POST.get('assessment_city')
            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')

            if billing_entity:
                pass
            else:
                billing_entity=0

            employees = []

            for i in range(1, 2):
                employees.append(request.POST.get('employee_id', ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'room_id': room_id, 'guesthouse_id': guesthouse_id,
                       'booking_datetime': booking_date,'city_id':city_id,'checkin_datetime':checkin_datetime+':00','checkout_datetime':checkout_datetime+':00',
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city,
                       'billing_entity_id': billing_entity, 'employees': employees,'no_of_nights':no_of_nights,
                       'reason_booking':reason_for_booking,'no_of_seats':1,'is_sms':1,'is_email':1}
            print(payload)
            url_taxi_booking = settings.API_BASE_URL + "add_guesthouse_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Employee/guesthouse-bookings/2", {})
            else:
                messages.error(request, 'Failed To Add Guesthouse Booking..!')
                return HttpResponseRedirect("/Corporate/Employee/guesthouse-bookings/2", {})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'employee_login_type' in request.session:
            request = get_request()
            login_type = request.session['employee_login_type']
            access_token = request.session['employee_access_token']

            gid = request.POST.get('guesthouse_id', '')

            payload = {'corporate_id': request.user.corporate_id, 'guesthouse_id': gid}

            url_hotel_types = settings.API_BASE_URL + "get_guesthouse_details"
            hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
            guesthouse = hotel_types['Guesthouse']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities_ass = cities['AssCity']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            from_city = request.POST.get('from_city', '')
            check_in = request.POST.get('check_in_datetime', '')
            check_out = request.POST.get('check_out_datetime', '')
            gid = request.POST.get('guesthouse_id', '')
            rid = request.POST.get('room_id', '')

            if id:
                return render(request, 'Company/Employee/add_guesthouse_booking.html', {'cities_ass':cities_ass,'guesthouses':guesthouse,
               'assessments':ass_code,'cities':cities, 'corp_access':access,'from_city':from_city,'check_in':check_in,'check_out':check_out,'gid':gid,'rid':rid})
            else:
                return render(request, 'Company/Employee/add_guesthouse_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


def view_guesthouse_details(request,id):
    if 'employee_login_type' in request.session:
        request = get_request()
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        payload = {'corporate_id': request.user.corporate_id, 'guesthouse_id':id}

        url_hotel_types = settings.API_BASE_URL + "get_guesthouse_details"
        hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
        guesthouse = hotel_types['Guesthouse']

        if id:
            return render(request, 'Company/Employee/view_guesthouse_details.html', {'guesthouses':guesthouse})
        else:
            return render(request, 'Company/Employee/view_guesthouse_details.html', {})
    else:
        return HttpResponseRedirect("/login")


def download_guesthouse_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'employee_login_type' in request.session:
        login_type = request.session['employee_login_type']
        access_token = request.session['employee_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_guesthouse_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        
        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Guesthouse-Bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Guesthouse Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Guesthouse Code',
        'Guesthouse Name',
        'Room Number',
        'Room Name',
        'Room Code',
        'Preferred Hotel',

        'TaxiVaxi Status',
        'Employees Name',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"


        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"
        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            'Billing Entity',
            'Travel request Code',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_city_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            bk['guest_house_code'],
            bk['guest_house_name'],
            bk['room_number'],
            bk['room_name'],
            bk['room_code'],


            bk['preferred_hotel'],


            bk['status_cotrav'],
            passanger_list,

            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),



            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response




def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
