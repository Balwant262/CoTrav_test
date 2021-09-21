from django.conf import settings
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
import requests
import json
import string
import random
from datetime import date, datetime
from django_global_request.middleware import get_request
from django.contrib import messages
from openpyxl import Workbook
from Common.VIEW.Agent.agent_views import upload_visa_doc_get_path, file_upload_get_path, upload_frro_doc_get_path
from Common.models import Corporate_Login_Access_Token
from xml.etree import ElementTree
import base64
import traceback
import xml.sax
import dateutil.parser
import ast
from landing.flight_parse import FlightHandler2 , FlightDetailsHandler , FlightBookingHandler , UniversalRecordRetrieveRsp , UniversalRecordCancelRsp

CREDENTIALS = 'Universal API/uAPI3568924042-3768c943:YGPNdGwezaGRrzqZEzWbf4AeN'
CREDENTIALS_enc64 = base64.b64encode(bytes(CREDENTIALS, 'utf-8')).decode("ascii")
Provider = '1G'
RGETBRANCH = 'P7009927'



def logout_action(request):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user = Corporate_Login_Access_Token.objects.get(access_token=access_token)
        del request.session['admin_login_type']
        del request.session['admin_access_token']

        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        #logout(request)  # the user is now LogOut
        return redirect("/login")
    else:
        return redirect("/login")


def homepage(request):

    if 'admin_login_type' in request.session:

        user_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        print(request.user)
        print("Admin Home Page ")
        print(user_type)
        print(access_token)
        payload = {'admin_id':request.user.id, 'corporate_id':request.user.corporate_id}
        url = settings.API_BASE_URL + "admin_dashboard"

        data = getDataFromAPI(user_type, access_token, url, payload)
        dataDashboard = data['Dashboard']
        print(dataDashboard)
        return render(request, 'Company/Admin/home_page.html', {'user': request.user,'dataDashboard':dataDashboard})
    else:
        print("i m from home els part")
        return HttpResponseRedirect("/login")

def user_profile(request):
    return render(request, 'Company/Admin/user_profile.html', {'user': request.user})


def company_admins(request, id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            admins = company['Admins']
            return render(request, "Company/Admin/company_admins.html", {'admins': admins})
        else:
            return render(request, "Company/Admin/company_admins.html", {'admins': {}})
    else:
        return HttpResponseRedirect("/login")


def assessment_cities(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id;

        if request.method == 'POST':
            corporate_id = request.POST.get('corporate_id', '')
            city_name = request.POST.get('city_name', '')
            current_url = request.POST.get('current_url', '')
            city_id = request.POST.get('city_id', '')

            payload = {'corporate_id':corporate_id,'city_name':city_name,'city_id':city_id,'login_type':login_type,'user_id':user_id}

            if city_id:
                url = settings.API_BASE_URL + "update_assessment_cities"
                opr_msg = "Assessment City Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_assessment_cities"
                opr_msg = "Assessment City Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, company['error'])
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "assessment_cities"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            if company['success'] == 1:
                cities = company['Cities']
                companies = company1['Corporates']
                return render(request, "Company/Admin/assessment_cities.html", {'cities': cities,'companies':companies})
            else:
                return render(request, "Company/Admin/assessment_cities.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/login")


def assessment_codes(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id;
        if request.method == 'POST':
            current_url = request.POST.get('current_url', '')
            corporate_id = request.POST.get('corporate_id', '')
            assessment_code = request.POST.get('assessment_code', '')
            code_desc = request.POST.get('code_desc', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            service_from = request.POST.get('service_from', '')
            service_to = request.POST.get('service_to', '')

            code_id = request.POST.get('code_id', '')

            payload = {'corporate_id': corporate_id, 'assessment_code': assessment_code, 'code_desc': code_desc,'from_date':from_date,'to_date':to_date,
                       'login_type': login_type, 'user_id': user_id,'service_from':service_from,'service_to':service_to,'code_id':code_id}

            if code_id:
                url = settings.API_BASE_URL + "update_assessment_codes"
                opr_msg = "Assessment Code Updated Successfully"
            else:
                url = settings.API_BASE_URL + "add_assessment_codes"
                opr_msg = "Assessment Code Added Successfully"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, "Operation Fail")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "assessment_codes"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            if company['success'] == 1:
                codes = company['Codes']
                companies = company1['Corporates']
                return render(request, "Company/Admin/assessment_codes.html", {'codes': codes,'companies':companies})
            else:
                return render(request, "Company/Admin/assessment_codes.html", {'codes': {}})
    else:
        return HttpResponseRedirect("/login")


def delete_assessment_codes(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        code_id = request.POST.get('code_id')
        current_url = request.POST.get('current_url', '')

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            payload = {'code_id': code_id, 'user_id': user_id, 'login_type': login_type}
            url = settings.API_BASE_URL + "delete_assessment_codes"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Assessment Code Deleted Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Deleted Successfully"})
            else:
                messages.error(request, 'Failed to Delete Assessment Code..!')
                return HttpResponseRedirect(current_url, {'message': "Fails"})
        else:
            return HttpResponseRedirect("/login")


def delete_assessment_cities(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        city_id = request.POST.get('city_id')
        current_url = request.POST.get('current_url', '')

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            payload = {'city_id': city_id, 'user_id': user_id, 'login_type': login_type}
            url = settings.API_BASE_URL + "delete_assessment_cities"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Assessment City Deleted Successfully')
                return HttpResponseRedirect(current_url, {'message': "Deleted Successfully"})
            else:
                messages.error(request, 'Failed to Delete Assessment City..!')
                return HttpResponseRedirect(current_url, {'message': "Fails"})
        else:
            return HttpResponseRedirect("/login")


def company_billing_entities(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': id}

        company = getDataFromAPI(login_type, access_token, url, payload)
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)

        if company['success'] == 1:
            entities = company['Entitys']
            cities = cities["Cities"]
            return render(request, "Company/Admin/billing_entities.html",
                          {'billing_entities': entities, "cities": cities, })
        else:
            return render(request, "Company/Admin/billing_entities.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_rates(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            company_rates = company['Corporate_Retes']
            return render(request, "Company/Admin/company_rates.html", {'corporate_rates': company_rates})
        else:
            return render(request, "Company/Admin/company_rates.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_groups(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            groups = company['Groups']
            return render(request, "Company/Admin/groups.html", {'groups': groups})
        else:
            return render(request, "Company/Admin/groups.html", {'groups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_subgroups(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            url2 = settings.API_BASE_URL + "groups"
            subgroups = company['Subgroups']
            gr = getDataFromAPI(login_type, access_token, url2, payload)
            groups = gr['Groups']
            return render(request, "Company/Admin/subgroups.html", {'subgroups': subgroups, 'groups': groups})
        else:
            return render(request, "Company/Admin/subgroups.html", {'subgroups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_spocs(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            spocs = company['Spocs']

            url_subgroup = settings.API_BASE_URL + "subgroups"
            company_sub = getDataFromAPI(login_type, access_token, url_subgroup, payload)
            subgroups = company_sub['Subgroups']

            return render(request, "Company/Admin/spocs.html", {'spocs': spocs, 'subgroups':subgroups})
        else:
            return render(request, "Company/Admin/spocs.html", {'spocs': {}})
    else:
        return HttpResponseRedirect("/login")


def company_employees(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            employees = company['Employees']
            return render(request, "Company/Admin/employees.html", {'employees': employees})
        else:
            return render(request, "Company/Admin/employees.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/login")


def add_company_rate(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            url = settings.API_BASE_URL + "company_rates"
            payload = {'corporate_id': id}
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, 'Company Package Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Admin/company-rates/" + id, {'message': "Added Successfully"})
            else:
                messages.error(request, "Similar Package Already Exists..!")
                return HttpResponseRedirect("/Corporate/Admin/company-rates/" + id, {'message': "Operation Fails"})
        else:
            return HttpResponseRedirect("/login")
    else:
        return render(request, "Company/Admin/company_rate_add.html", {'entities': {}})
        pass


def add_company_entity(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            current_url = request.POST.get('current_url', '')
            entity_name = request.POST.get('entity_name', '')
            billing_city_id = request.POST.get('billing_city_id')
            contact_person_name = request.POST.get('contact_person_name', '')
            contact_person_email = request.POST.get('contact_person_email', '')
            contact_person_no = request.POST.get('contact_person_no', '')
            address_line_1 = request.POST.get('address_line_1', '')
            address_line_2 = request.POST.get('address_line_2', '')
            address_line_3 = request.POST.get('address_line_3', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            entity_id = request.POST.get('entity_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token,
                       'entity_name': entity_name, 'billing_city_id': billing_city_id,
                       'contact_person_name': contact_person_name, 'contact_person_email': contact_person_email,
                       'contact_person_no': contact_person_no, 'address_line_1': address_line_1,
                       'address_line_2': address_line_2,
                       'address_line_3': address_line_3, 'gst_id': gst_id, 'pan_no': pan_no, 'entity_id': entity_id,
                       'is_delete': delete_id, }

            if entity_id:
                url = settings.API_BASE_URL + "update_billing_entity"
                operation_message = "Company Entity Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_billing_entity"
                    operation_message = "Company Entity Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_billing_entity"
                operation_message = "Company Entity Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url,{})
            else:
                messages.error(request, 'Billing Entity Already Exists..!')
                return HttpResponseRedirect(current_url,{})
        else:
            return HttpResponseRedirect("/login")


def add_company_group(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = id
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            current_url = request.POST.get('current_url', '')
            group_name = request.POST.get('group_name', '')
            zone_name = request.POST.get('zone_name')

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_visa = request.POST.get('is_visa', '')
            is_frro = request.POST.get('is_frro', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))
            password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'group_name': group_name, 'zone_name': zone_name,'access_token_auth':access_token_auth,'name':name,
                       'email':email,'cid':cid,'contact_no':contact_no,'is_radio':is_radio,'is_local':is_local,'is_outstation':is_outstation,'is_bus':is_bus,
                       'is_train':is_train,'is_hotel':is_hotel,'is_meal':is_meal,'is_flight':is_flight,'is_water_bottles':is_water_bottles,'is_reverse_logistics':is_reverse_logistics,
                       'password':password,'is_send_email':is_send_email,'is_send_sms':is_send_sms,'is_visa':is_visa,'is_frro':is_frro,'is_guesthouse':is_guesthouse}

            url = settings.API_BASE_URL + "add_group"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Company Group Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, 'Group Name Already Exists..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()
        if 'admin_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            current_url = request.POST.get('current_url', '')
            subgroup_name = request.POST.get('group_name', '')
            group_id = request.POST.get('group_id', '')

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_visa = request.POST.get('is_visa', '')
            is_frro = request.POST.get('is_frro', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))
            password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'subgroup_name': subgroup_name, 'group_id': group_id,'name':name,
                       'email':email,'cid':cid,'contact_no':contact_no,'is_radio':is_radio,'is_local':is_local,'is_outstation':is_outstation,'is_bus':is_bus,
                       'is_train':is_train,'is_hotel':is_hotel,'is_meal':is_meal,'is_flight':is_flight,'is_water_bottles':is_water_bottles,'is_reverse_logistics':is_reverse_logistics,
                       'password':password,'is_send_email':is_send_email,'is_send_sms':is_send_sms,'is_visa':is_visa,'is_frro':is_frro,'is_guesthouse':is_guesthouse}

            #print(payload)
            url = settings.API_BASE_URL + "add_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print("sdsadasd")
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Add Company Subgroup Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, 'Subgroup Name Already Exists..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def update_company_group(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            group_id = request.POST.get('group_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            group_name = request.POST.get('group_name', '')
            zone_name = request.POST.get('zone_name')

            payload = {'group_id': group_id, 'access_token': access_token, 'group_name': group_name, 'zone_name': zone_name,
                       'user_id': user_id, 'login_type': login_type}

            print(payload)
            url = settings.API_BASE_URL + "update_group"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Group Updated Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Update Successfully"})
            else:
                messages.error(request, 'Company Group Update Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Updated"})
        else:
            return HttpResponseRedirect("/login")


def update_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            group_name = request.POST.get('group_name', '')

            payload = {'subgroup_id': subgroup_id, 'access_token': access_token, 'group_name': group_name,
                       'user_id': user_id, 'login_type': login_type}

            print(payload)
            url = settings.API_BASE_URL + "update_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Subgroup Updated Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Update Successfully"})
            else:
                messages.error(request, 'Company Subgroup Update Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Updated"})
        else:
            return HttpResponseRedirect("/login")


def delete_company_group(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            group_id = request.POST.get('group_id')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            access_token_auth = request.session['admin_access_token']
            payload = {'group_id': group_id, 'user_id': user_id, 'login_type': login_type, 'access_token': access_token,
                       'access_token_auth': access_token_auth}
            url = settings.API_BASE_URL + "delete_group"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Group Deleted Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Delete Successfully"})
            else:
                messages.error(request, 'Company Group Deletion Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def delete_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            subgroup_id = request.POST.get('subgroup_id')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = request.session['admin_access_token']

            payload = {'subgroup_id': subgroup_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'access_token_auth': access_token_auth}
            url = settings.API_BASE_URL + "delete_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(payload)
            if company['success'] == 1:
                messages.success(request, 'Company Sub-group Deleted Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Delete Successfully"})
            else:
                messages.error(request, 'Company Sub-group Deletion Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_group_auth(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            group_id = request.POST.get('group_id')
            group_auth_id = request.POST.get('group_auth_id')
            delete_id = request.POST.get('delete_id')
            is_frro = request.POST.get('is_frro', '')
            is_visa = request.POST.get('is_visa', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')

            if group_id:
                group_auth_id = group_auth_id
                password = "taxi123"

            if group_auth_id:
                group_auth_id = group_auth_id
            else:
                group_auth_id = 0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'group_id': group_id, 'delete_id': delete_id, 'password': password, 'group_auth_id': group_auth_id,
                       'access_token_auth': access_token_auth,'is_send_sms':is_send_sms,'is_send_email':is_send_email,'is_visa':is_visa,
                       'is_frro':is_frro,'is_guesthouse':is_guesthouse}

            url = ""
            if group_auth_id:
                url = settings.API_BASE_URL + "update_group_auth"
                oper_msg = "Company Group Authentication Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_group_auth"
                    oper_msg = "Company Group Authentication Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_group_auth"
                oper_msg = "Company Group Authentication Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, oper_msg)
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Authenticator Already Exists. Try with another data..!")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_subgroup_auth(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            subgroup_id = request.POST.get('subgroup_id')
            subgroup_auth_id = request.POST.get('subgroup_auth_id')
            delete_id = request.POST.get('delete_id')
            is_frro = request.POST.get('is_frro', '')
            is_visa = request.POST.get('is_visa', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')

            if subgroup_id:
                subgroup_auth_id = subgroup_auth_id
                password = "taxi123"

            if subgroup_auth_id:
                subgroup_auth_id = subgroup_auth_id
            else:
                subgroup_auth_id = 0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'subgroup_id': subgroup_id, 'delete_id': delete_id, 'password': password,'is_visa':is_visa,'is_frro':is_frro,
                       'subgroup_auth_id': subgroup_auth_id, 'access_token_auth': access_token_auth,'is_send_email':is_send_email,
                       'is_send_sms':is_send_sms,'is_guesthouse':is_guesthouse}

            url = ""
            if subgroup_auth_id:
                url = settings.API_BASE_URL + "update_subgroup_auth"
                operation_msg = "Company SubGroup Authentication Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_subgroup_auth"
                    operation_msg = "Company SubGroup Authentication Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_subgroup_auth"
                operation_msg = "Company SubGroup Authentication Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_msg)
                return HttpResponseRedirect(current_url,{'message': "Added Successfully"})
            else:
                messages.error(request, "Authenticator Already Exists. Try with another data..!")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_admins(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            has_billing_access = request.POST.get('has_billing_access', '')
            admin_id = request.POST.get('admin_id')
            is_frro = request.POST.get('is_frro', '')
            is_visa = request.POST.get('is_visa', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')

            delete_id = request.POST.get('delete_id')

            if admin_id:
                password = ''
            else:
                password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,'is_guesthouse':is_guesthouse,
                       'admin_id': admin_id, 'delete_id': delete_id, 'password': password,'has_billing_access':has_billing_access,
                       'access_token_auth': access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms,'is_frro':is_frro,'is_visa':is_visa}

            url = ""
            print(payload)
            if admin_id:
                url = settings.API_BASE_URL + "update_admin"
                opration_msg = "Company Admin Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_admin"
                    opration_msg = "Company Admin Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_admin"
                opration_msg = "Company Admin Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, opration_msg)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, "Admin Already Exists. Try with another data..!")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_spocs(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')
            user_cid = request.POST.get('user_cid', '')

            user_name = request.POST.get('user_name', '')
            user_contact = request.POST.get('user_contact', '')
            email = request.POST.get('email', '')
            username = request.POST.get('email', '')
            budget = request.POST.get('budget', '')
            expense = request.POST.get('budget', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            spoc_id = request.POST.get('spoc_id')
            is_frro = request.POST.get('is_frro', '')
            is_visa = request.POST.get('is_visa', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')

            delete_id = request.POST.get('delete_id')

            if spoc_id:
                password = ''
            else:
                password = "taxi123"
                spoc_id =0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'group_id': group_id, 'subgroup_id': subgroup_id, 'user_cid': user_cid, 'user_name': user_name,
                       'user_contact':user_contact,'email':email,'username':username,'budget':budget,'expense':expense,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'spoc_id': spoc_id, 'delete_id': delete_id, 'password': password,'is_send_sms':is_send_sms,'is_send_email':is_send_email,
                       'is_frro':is_frro,'is_visa':is_visa,'is_guesthouse':is_guesthouse}

            url = ""
            print(payload)
            if spoc_id:
                url = settings.API_BASE_URL + "update_spoc"
                success_msg = "Spoc Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_spoc"
                    success_msg = "Spoc De-Activated Successfully..!"
                if delete_id == '2':
                    url = settings.API_BASE_URL + "active_spoc"
                    success_msg = "Spoc Activated Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_spoc"
                success_msg = "Spoc Added Successfully..1"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, success_msg)
                return HttpResponseRedirect("/Corporate/Admin/company-spoc/"+str(corporate_id), {'message': "Added Successfully"})
            else:
                messages.error(request, "Spoc Already Exists. Try with another data..!")
                return HttpResponseRedirect("/Corporate/Admin/company-spoc/"+str(corporate_id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")
    else:
        request = get_request()

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_spoc = settings.API_BASE_URL + "view_spoc"
            payload = {'spoc_id': id}
            r = requests.post(url_spoc, data=payload, headers=headers)
            company_spoc = json.loads(r.text)
            spocs = company_spoc['Spoc']

            url = settings.API_BASE_URL + "groups"
            payload = {'corporate_id': request.user.corporate_id}
            company = getDataFromAPI(login_type, access_token, url, payload)
            groups = company['Groups']

            url_subgroup = settings.API_BASE_URL + "subgroups"
            company_sub = getDataFromAPI(login_type, access_token, url_subgroup, payload)
            subgroups = company_sub['Subgroups']

            if id:
                return render(request, 'Company/Admin/add_spoc.html', {'groups': groups, 'subgroups': subgroups, 'spoc':spocs})
            else:
                return render(request, 'Company/Admin/add_spoc.html', {'groups': groups, 'subgroups': subgroups})
        else:
            return HttpResponseRedirect("/login")


def add_employee(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            spoc_id = request.POST.get('spoc_id', '')
            billing_entity_id = request.POST.get('billing_entity_id', '')
            core_employee_id = request.POST.get('core_employee_id', '')
            employee_cid = request.POST.get('employee_cid', '')

            employee_name = request.POST.get('employee_name', '')
            employee_email = request.POST.get('employee_email', '')
            username = request.POST.get('employee_email', '')
            employee_contact = request.POST.get('employee_contact', '')

            date_of_birth = request.POST.get('date_of_birth','')
            if date_of_birth and date_of_birth != 'None':
                age = calculate_age(date_of_birth)
            else:
                age = 0

            gender = request.POST.get('gender')
            id_proof_type = request.POST.get('id_proof_type')

            id_proof_no = request.POST.get('id_proof_no', '')
            is_active = request.POST.get('is_active', '')
            has_dummy_email = 0
            fcm_regid = request.POST.get('fcm_regid', '')
            is_cxo = request.POST.get('is_cxo', '')
            designation = request.POST.get('designation', '')
            home_city = request.POST.get('home_city', '')
            home_address = request.POST.get('home_address', '')
            reporting_manager = request.POST.get('reporting_manager', '')
            employee_band = request.POST.get('employee_band', '')

            is_frro = request.POST.get('is_frro', '')
            is_visa = request.POST.get('is_visa', '')
            is_guesthouse = request.POST.get('is_guesthouse', '')
            passport_issue_date = request.POST.get('passport_issue_date', '')
            passport_no = request.POST.get('passport_no', '')
            passport_expiry_date = request.POST.get('passport_expiry_date', '')
            visa_number = request.POST.get('visa_number', '')
            visa_expiry_date = request.POST.get('visa_expiry_date', '')
            current_location = request.POST.get('current_location', '')
            current_stay_address = request.POST.get('current_stay_address', '')
            company_contract_expiry_date = request.POST.get('company_contract_expiry_date', '')
            lease_status = request.POST.get('lease_status', '')
            
            if is_cxo == '1':
                assistant_id = request.POST.get('assistant_id', '')
                if assistant_id == ' ':
                    assistant_id =0
            else:
                assistant_id = 0

            employee_id = request.POST.get('employee_id')

            delete_id = request.POST.get('delete_id')

            if employee_id:
                password = ''
            else:
                password = "taxi123"
                employee_id =0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,'spoc_id':spoc_id,'core_employee_id':core_employee_id,
                       'access_token': access_token,'employee_cid':employee_cid,'employee_name':employee_name,'employee_email':employee_email,'is_guesthouse':is_guesthouse,
                       'employee_contact':employee_contact,'age':age,'gender':gender,'id_proof_type':id_proof_type,'id_proof_no':id_proof_no,
                       'is_active':is_active,'has_dummy_email':has_dummy_email,'fcm_regid':fcm_regid,'is_cxo':is_cxo,'employee_id': employee_id,
                       'designation':designation,'home_city':home_city,'home_address':home_address,'assistant_id':assistant_id,'date_of_birth':date_of_birth,
                       'delete_id': delete_id, 'password': password,'billing_entity_id':billing_entity_id,'username':username,'reporting_manager':reporting_manager,
                       'employee_band':employee_band,'is_frro':is_frro,'is_visa':is_visa,'passport_issue_date':passport_issue_date,'passport_no':passport_no,
                       'passport_expiry_date':passport_expiry_date,'visa_number':visa_number,'visa_expiry_date':visa_expiry_date,'current_location':current_location,
                       'current_stay_address':current_stay_address,'company_contract_expiry_date':company_contract_expiry_date,'lease_status':lease_status}

            url = ""
            
            if employee_id:
                url = settings.API_BASE_URL + "update_employee"
                success_msg  = "Employee Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_employee"
                    success_msg  = "Employee Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_employee"
                success_msg = "Employee Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            
            if company['success'] == 1:
                messages.success(request, success_msg)
                return HttpResponseRedirect("/Corporate/Admin/company-employees/"+str(corporate_id), {'message': "Added Successfully"})
            else:
                messages.error(request, "Employee Already Exists. Try with another data..!")
                return HttpResponseRedirect("/Corporate/Admin/company-employees/"+str(corporate_id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")
    else:
        request = get_request()

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            url_emp = settings.API_BASE_URL + "view_employee"
            payload = {'employee_id': id, 'corporate_id' : request.user.corporate_id}

            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employee']

            url_spoc = settings.API_BASE_URL + "spocs"
            payload_spoc = {'corporate_id': request.user.corporate_id}
            company_spoc = getDataFromAPI(login_type, access_token, url_spoc, payload_spoc)
            spocs = company_spoc['Spocs']

            url_entity = settings.API_BASE_URL + "billing_entities"
            payload_entity = {'corporate_id': request.user.corporate_id}
            company_entity = getDataFromAPI(login_type, access_token, url_entity, payload_entity)
            entitys1 = company_entity['Entitys']

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']

            url_emp = settings.API_BASE_URL + "employees"
            employees1 = getDataFromAPI(login_type, access_token, url_emp, payload)
            employeess = employees1['Employees']

            if id:
                return render(request, 'Company/Admin/add_employee.html', {'employee':employees,'spocs':spocs,'entitys':entitys1,'cities':cities,'employees':employeess})
            else:
                return render(request, 'Company/Admin/add_employee.html', {'spocs':spocs,'entitys':entitys1,'cities':cities,'employees':employeess})
        else:
            return HttpResponseRedirect("/login")


def view_company_group(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_group"
        payload = {'group_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_group_auth"
        grp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(grp_auths)
        if company['success'] == 1:
            groups = company['Groups']
            grp_auths = grp_auths['Groups']
            return render(request, "Company/Admin/view_groups.html", {'group': groups, 'grp_auths': grp_auths})
        else:
            return render(request, "Company/Admin/view_groups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")


def view_company_subgroup(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_subgroup"
        payload = {'subgroup_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_subgroup_auth"
        subgrp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(subgrp_auths)
        if company['success'] == 1:
            subgroups = company['SubGroups']
            subgrp_auths = subgrp_auths['SubGroups']
            return render(request, "Company/Admin/view_subgroups.html",
                          {'subgroup': subgroups, 'subgrp_auths': subgrp_auths})
        else:
            return render(request, "Company/Admin/view_subgroups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")

# #####################   TAXI ###########################################

def taxi_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_taxi_bookings"
        payload = {'booking_type': id,'corporate_id':corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/taxi_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/taxi_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_taxi_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_taxi_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_taxi_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_taxi_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')
        user_id = request.user.id

        url = settings.API_BASE_URL + "admin_accept_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Taxi Booking Failed To Accept..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_taxi_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            messages.error(request, 'Taxi Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Taxi Booking')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_taxi_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            user_id = request.POST.get('user_id', '')

            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            tour_type = request.POST.get('tour_type', '')
            pickup_city = request.POST.get('pickup_city', '')
            pickup_location = request.POST.get('pickup_location', '')
            drop_location = request.POST.get('drop_location', '')
            pickup_datetime = request.POST.get('pickup_datetime', '')
            taxi_type = request.POST.get('taxi_type', '')
            package_id = request.POST.get('package_id', '')
            no_of_days = request.POST.get('no_of_days', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            actual_city_id = request.POST.get('current_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            if tour_type == 1 or tour_type == '1':
                url_add_city = settings.API_BASE_URL + "add_city_name"
                pickup_details = [x.strip() for x in pickup_city.split(',')]
                city_data = {'login_type': login_type, 'access_token': access_token, 'city_name': pickup_details[0], 'state_id': '1'}
                city_id = getDataFromAPI(login_type, access_token, url_add_city, city_data)
                for conty_id in city_id['id']:
                    actual_city_id = conty_id['id']

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'tour_type':tour_type,'pickup_city':actual_city_id,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'pickup_location':pickup_location,'drop_location':drop_location,'pickup_datetime':pickup_datetime+':00','taxi_type':taxi_type,
                       'package_id':package_id,'no_of_days':no_of_days,'reason_booking':reason_booking,'no_of_seats':no_of_seats,
                       'employees':employees,'user_id':user_id,'entity_id':entity_id,'is_sms':1,'is_email':1}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_taxi_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/taxi-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed To Add Taxi Booking')
                return HttpResponseRedirect("/Corporate/Admin/taxi-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_taxi = settings.API_BASE_URL + "taxi_types"
            taxies = getDataFromAPI(login_type, access_token, url_taxi, payload)
            taxies = taxies['taxi_types']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_city1 = settings.API_BASE_URL + "city_by_package"
            cities1 = getDataFromAPI(login_type, access_token, url_city1, payload)
            citiess = cities1['Cities']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_taxi_booking.html', {'employees':employees,'entities':entities,'cities':cities,
                'taxies':taxies,'assessments':ass_code,'citiess':citiess, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_taxi_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

#####################################  BUS  #####################################


def bus_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_bus_bookings"
        payload = {'corporate_id': corporate_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/bus_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_bus_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_bus_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_bus_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Accepted Successfully...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Bus Booking...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_bus_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Bus Booking...!')
            return HttpResponseRedirect(current_url , {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_bus_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            bus_type = request.POST.get('bus_type', '')
            bus_type2 = request.POST.get('bus_type2', '')
            bus_type3 = request.POST.get('bus_type3', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            preferred_board_point = request.POST.get('preferred_board_point', '')
            preferred_drop_point = request.POST.get('preferred_drop_point', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'bus_type':bus_type,'booking_datetime':booking_datetime,'journey_datetime':journey_datetime+':00','journey_datetime_to':journey_datetime_to+':00','entity_id':entity_id,
                       'preferred_bus':preferred_bus,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'is_sms':1,'is_email':1,
                       'preferred_board_point':preferred_board_point, 'preferred_drop_point':preferred_drop_point,'bus_type2':bus_type2,'bus_type3':bus_type3}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_bus_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/bus-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Bus Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/bus-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_bus_type = settings.API_BASE_URL + "bus_types"
            bus_type = getDataFromAPI(login_type, access_token, url_bus_type, payload)
            bus_types = bus_type['Types']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_bus_booking.html', {'bus_types':bus_types,'employees':employees,'cities':cities,
                 'entities':entities,'assessments':ass_code, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_bus_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

#####################################  TRAIN  #####################################


def train_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_train_bookings"
        payload = {'corporate_id':corporate_id,'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/train_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_train_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_train_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_train_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Accepted Successfully...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Train Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_train_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Train Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_train_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            train_type = request.POST.get('train_type', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'train_type':train_type,'booking_datetime':booking_datetime,'journey_datetime':journey_datetime+':00','entity_id':entity_id,
                       'preferred_bus':preferred_bus,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,
                       'is_sms':1,'is_email':1,'journey_datetime_to':journey_datetime_to+':00'}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_train_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/train-bookings/1", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Train Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/train-bookings/1", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_bus_type = settings.API_BASE_URL + "train_types"
            bus_type = getDataFromAPI(login_type, access_token, url_bus_type, payload)
            train_types = bus_type['Types']


            # url_railway_stations = settings.API_BASE_URL + "railway_stations"
            # trains1 = getDataFromAPI(login_type, access_token, url_railway_stations, payload)
            # railway_stations = trains1['Stations']
            railway_stations = ""

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_train_booking.html', {'train_types':train_types,'employees':employees,'cities':cities,
                'entities':entities,'assessments':ass_code,'railway_stations':railway_stations, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_train_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


#####################################  Hotels  #####################################


def hotel_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_hotel_bookings"
        payload = {'corporate_id':corporate_id,'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/hotel_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_hotel_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_hotel_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_hotel_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Hotel Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_hotel_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Hotel Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_hotel_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_city = request.POST.get('from_city')
            city_area = request.POST.get('city_area')
            preferred_hotel_area = request.POST.get('preferred_hotel_area')
            check_in_date = request.POST.get('check_in_date')
            check_out_date = request.POST.get('check_out_date')
            room_type_priority1 = request.POST.get('room_type_priority1')
            room_type_priority2 = request.POST.get('room_type_priority2')
            room_occupancy = request.POST.get('room_occupancy')
            preferred_hotel = request.POST.get('preferred_hotel')
            booking_date = request.POST.get('booking_datetime')
            no_of_nights = request.POST.get('no_of_nights', '')
            assessment_code = request.POST.get('assessment_code')

            assessment_city = request.POST.get('assessment_city')
            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')

            no_of_seats = 1

            employees = []

            for i in range(1,2):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from_city_id': from_city, 'from_area_id': city_area,
                       'preferred_area': preferred_hotel_area, 'checkin_datetime': check_in_date+':00',
                       'checkout_datetime': check_out_date+':00', 'bucket_priority_1': room_type_priority1,
                       'bucket_priority_2': room_type_priority2, 'room_type_id': room_occupancy,
                       'preferred_hotel': preferred_hotel, 'booking_datetime': booking_date,'no_of_nights':no_of_nights,
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city,
                       'billing_entity_id': billing_entity, 'employees': employees,'reason_booking':reason_for_booking,'no_of_seats':no_of_seats,
                       'is_sms':1,'is_email':1}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_hotel_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/hotel-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Hotel Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/hotel-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            # url_city = settings.API_BASE_URL + "cities"
            # cities = getDataFromAPI(login_type, access_token, url_city, payload)
            # cities = cities['Cities']
            cities = ""

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_room_types = settings.API_BASE_URL + "room_types"
            room_types = getDataFromAPI(login_type, access_token, url_room_types, payload)
            room_types = room_types['Types']

            url_hotel_types = settings.API_BASE_URL + "hotel_types"
            hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
            hotel_types = hotel_types['Types']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_hotel_booking.html', {'hotel_types':hotel_types,'room_types':room_types,'employees':employees,
                'cities':cities,'entities':entities,'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_hotel_booking.html', {})
        else:
            return HttpResponseRedirect("/login")
#####################################  FLIGHT  #####################################


def flight_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_flight_bookings"
        payload = {'corporate_id':corporate_id,'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/flight_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/flight_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_flight_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_flight_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_flight_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_flight_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Accepted Successfully...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Flight Booking...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_flight_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Flight Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_flight_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')

            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            usage_type = request.POST.get('usage_type', '')
            trip_type = request.POST.get('trip_type', '')
            seat_type = request.POST.get('seat_type', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            departure_date = request.POST.get('departure_date', '')
            preferred_flight = request.POST.get('preferred_flight', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            if entity_id:
                pass
            else:
                entity_id=0

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'usage_type':usage_type,'trip_type':trip_type,'seat_type':seat_type,'from_city':from_city,'to_city':to_city,
                       'booking_datetime':booking_datetime,'departure_datetime':departure_date,'preferred_flight':preferred_flight,'assessment_code':assessment_code,
                       'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'billing_entity_id':entity_id,
                       'is_sms':1,'is_email':1,'assessment_city_id':assessment_city_id}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
            print(booking)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Flight Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            if id:
                return render(request, 'Company/Admin/add_flight_booking.html', {'employees':employees,'entities':entities,
                            'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs,'airports':airports})
            else:
                return render(request, 'Company/Admin/add_flight_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


def add_self_flight_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('admin_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            trip_type = request.POST.get('trip_type', '')
            return_date = request.POST.get('return_date', '')
            fl_class = request.POST.get('fl_class', '')
            no_of_seats = request.POST.get('no_of_seats', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            departure_date = request.POST.get('departure_date', '')
            dep_data = dateutil.parser.parse(departure_date).strftime('%Y-%m-%d')

            return_date_data = ''
            if trip_type == '2' or trip_type == 2:
                return_data = dateutil.parser.parse(return_date).strftime('%Y-%m-%d')
                return_date_data = """
                <air:SearchAirLeg>
                    <air:SearchOrigin>
                      <com:CityOrAirport Code="""+'"'+to_city+'"'+""" />
                    </air:SearchOrigin>
                    <air:SearchDestination>
                      <com:CityOrAirport Code="""+'"'+from_city+'"'+""" />
                    </air:SearchDestination>
                    <air:SearchDepTime PreferredTime="""+'"'+return_data+'"'+""" />
                    <air:AirLegModifiers>
                      <air:PermittedCabins>
                        <CabinClass Type="""+'"'+fl_class+'"'+""" xmlns="http://www.travelport.com/schema/common_v50_0" />
                      </air:PermittedCabins>
                    </air:AirLegModifiers>
                  </air:SearchAirLeg>
                """
            no_of_pass_str = '<com:SearchPassenger Code="ADT" />'
            n = int(no_of_seats)-1
            for i in range(n):
                no_of_pass_str += '\n' + no_of_pass_str

            booking_data = {'user_id': user_id, 'user_type': login_type, 'corporate_id': corporate_id,
                            'spoc_id': spoc_id, 'group_id': group_id,
                            'subgroup_id': subgroup_id, 'from_city': from_city, 'to_city': to_city,
                            'departure_datetime': departure_date, 'return_date': return_date, 'trip_type': trip_type,
                            'fl_class': fl_class, 'no_of_seats': no_of_seats}

            payload = {'auth_token': "", 'session_id': access_token, 'from_city': from_city, 'to_city': to_city, 'departure_date': departure_date,
                       'fl_class': fl_class, 'return_date': return_date, 'trip_type': trip_type, 'no_of_seats': no_of_seats }
            print(payload)
            try:
                payload = """
                        <soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/" >
                          <soap:Body>
                            <air:LowFareSearchReq TargetBranch="""+'"'+RGETBRANCH+'"'+""" TraceId="PP_1G_001322" SolutionResult="false" AuthorizedBy="SUSIL" xmlns:air="http://www.travelport.com/schema/air_v50_0" xmlns:com="http://www.travelport.com/schema/common_v50_0">
                              <com:BillingPointOfSaleInfo OriginApplication="UAPI" />
                              <air:SearchAirLeg>
                                <air:SearchOrigin>
                                  <com:Airport Code="""+'"'+from_city+'"'+""" />
                                </air:SearchOrigin>
                                <air:SearchDestination>
                                  <com:Airport Code="""+'"'+to_city+'"'+""" />
                                </air:SearchDestination>
                                <air:SearchDepTime PreferredTime="""+'"'+dep_data+'"'+""" />
                              </air:SearchAirLeg>
                              
                              """+return_date_data+"""
                              
                               <air:AirSearchModifiers>
                                <air:PreferredProviders>
                                  <com:Provider Code="""+'"'+Provider+'"'+""" />
                                </air:PreferredProviders>
                                <air:PermittedCabins>
                                  <CabinClass Type="""+'"'+fl_class+'"'+""" xmlns="http://www.travelport.com/schema/common_v50_0" />
                                </air:PermittedCabins>
                              </air:AirSearchModifiers>

                              """+no_of_pass_str+"""
                            </air:LowFareSearchReq>
                          </soap:Body>
                        </soap:Envelope>
                      """
                header = {
                    "Content-Type": "text/xml:charset=utf-8",
                    "Accept": "gzip,deflate",
                    "Connection": "Keep-Alive",
                    "Authorization": "Basic %s" % CREDENTIALS_enc64,
                    "Content-Length": str(len(payload))
                }

                url = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/AirService"
                print(payload)
                response = requests.post(url, data=payload, headers=header)
                flightdata = ElementTree.fromstring(response.content)
                url_access = settings.API_BASE_URL + "get_airports"
                data11 = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data11['Airports']

                return render(request, 'Company/Admin/add_flight_booking_serarch_result.html', {'booking_datas': booking_data, 'airports': airports, 'flights':flightdata})

            except Exception as e:
                print("EXCEPTIONNNNNNNNNNNN")
                print(traceback.format_exc())
                url_access = settings.API_BASE_URL + "get_airports"
                data11 = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data11['Airports']
                messages.success(request, 'No Flight Found Please Try Another Flight.!')
                return render(request, 'Company/Admin/add_self_flight_booking.html',
                              {'booking_datas': booking_data, 'flights': '', 'airports': airports})
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            if id:
                return render(request, 'Company/Admin/add_self_flight_booking.html', {'employees':employees,'entities':entities,
                            'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs,'airports':airports})
            else:
                return render(request, 'Company/Admin/add_self_flight_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


def add_self_flight_booking_conformation(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        segmentdata = request.POST.get('segmentdata', '')
        segmentdata = ast.literal_eval(segmentdata)
        string = "<air:AirSegment "

        for k, v in segmentdata.items():
            print("%s: %s" % (k, v))
            string += str(k) + "=" + '"' + str(v) + '"' + " "
        string += 'NumberOfStops="1" ProviderCode="'+Provider+'" />'

        payload = """
                   <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/">
         <soapenv:Header/>
         <soapenv:Body>
          <air:AirPriceReq xmlns:com="http://www.travelport.com/schema/common_v50_0" xmlns:air="http://www.travelport.com/schema/air_v50_0" TargetBranch="P7038885" CheckOBFees="All" AuthorizedBy="user" TraceId="P7038885">
           <com:BillingPointOfSaleInfo OriginApplication="UAPI"/>
           <air:AirItinerary>

             """+string+"""

           </air:AirItinerary>
           <air:AirPricingModifiers FaresIndicator="AllFares">

           </air:AirPricingModifiers>
           <com:SearchPassenger Code="ADT" Key="1" Age="50"/>

           <air:AirPricingCommand/>
          </air:AirPriceReq>
         </soapenv:Body>
        </soapenv:Envelope>
                                """

        url = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/AirService"

        header = {
            "Content-Type": "text/xml:charset=utf-8",
            "Accept": "gzip,deflate",
            "Connection": "Keep-Alive",
            "Authorization": "Basic %s" %CREDENTIALS_enc64,
            "Content-Length": str(len(payload))
        }

        response = requests.post(url, data=payload, headers=header)
        flightdata = ElementTree.fromstring(response.content)

        payload = {'corporate_id': request.user.corporate_id, 'spoc_id': request.user.id}
        url_enty = settings.API_BASE_URL + "billing_entities"
        entys = getDataFromAPI(login_type, access_token, url_enty, payload)
        entities = entys['Entitys']

        url_ass_code = settings.API_BASE_URL + "get_assessment_code"
        ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
        ass_code = ass_code['AssCodes']

        url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
        cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
        cities = cities_ass['AssCity']

        url_emp = settings.API_BASE_URL + "employees"
        company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
        employees = company_emp['Employees']

        url_nat = settings.API_BASE_URL + "get_nationality"
        nationality = getDataFromAPI(login_type, access_token, url_nat, payload)
        nationalities = nationality['Nationality']

        return render(request, 'Company/Admin/add_flight_booking_conformation.html', {'employees': employees,
        'cities_ass': cities,'entities': entities, 'assessments': ass_code, 'nationalities':nationalities, 'flightdata':flightdata})

    else:
        return HttpResponseRedirect("/login")
####################### Download MIS ##################################

def dateonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError:
        return ''


def download_taxi_bookings(request):
    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

        # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [

        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',

        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'Passengers',

        'Booking Date',
        'Booking Time',

        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        'Approver1 Action',

        'Approver1 Name',

        'Approver1 Date',

        'Approver1 Time',

        'Approver2 Action',

        'Approver2 Name',

        'Approver2 Date',

        'Approver2 Time',

        'Approved Date',

        'Approved Time',

        'Approved By',

        'Canceled Date',

        'Canceled Time',

        'Canceled By',

        'Assigned Date',

        'Assigned Time',

        'Assigned By',

        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',

        'Package Name',
        'Tour Type',
        'Vehicle Type',

        'Driver Name',
        'Driver Contact	',
        'Taxi Reg No.',
        'No. Of Seats',

        'Current Booking Status',

        'Hours Done',

        'Allowed Hours',

        'Extra Hours',

        'Kms Done',

        'Allowed Kms',

        'Extra Kms',

        'Extra Hours Charges',

        'Base Price',

        'Management Fee',
        'Tax on management fee',
        'Sub Total',
        'Cotrav Billing Entity',
        'IGST',
        'CGST',
        'SGST',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',

        'Estimated Amount',
        'Is Auto Approved',
        'Bill ID',
        'Bill Date',
        'Billing Entity',

        'Is Auto Approved',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        spoc_status =''
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by= ''
        spoc_canceled_date =''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

            if bk['spoc_status'] == 1:
                spoc_status = "In-Active"
            else:
                spoc_status = "Active"

            if (act['action'] == 3 and act['user_type'] == 6):
                print('canceled')
                spoc_canceled_by = act['employee_name']
                spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [
            bk['reference_no'],
            bk['assessment_city_id'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],

            bk['zone_name'],
            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),
            dateonly(bk['pickup_datetime']),
            timeonly(bk['pickup_datetime']),

            bk['package_name'],
            bk['tour_type'],
            bk['taxi_type_name'],

            bk['driver_name'],
            bk['driver_contact'],
            'Taxi Reg No.',
            bk['no_of_seats'],

            bk['cotrav_status'],

            bk['ci_hours_done'],

            bk['ci_allowed_hours'],

            bk['ci_extra_hours'],

            bk['ci_kms_done'],

            bk['ci_allowed_kms'],

            bk['ci_extra_kms'],

            bk['ci_extra_hr_charges'],

            bk['base_rate'],

            bk['ci_management_fee'],
            bk['ci_tax_on_management_fee'],
            bk['ci_sub_total'],
            bk['ci_cotrav_billing_entity'],
            bk['ci_igst'],
            bk['ci_cgst'],
            bk['ci_sgst'],
            bk['ci_management_fee_igst'],
            bk['ci_management_fee_cgst'],
            bk['ci_management_fee_sgst'],
            bk['ci_management_fee_igst_rate'],
            bk['ci_management_fee_cgst_rate'],
            bk['ci_management_fee_sgst_rate'],

            '',
            '',
            '',
            '',
            '',
            '',

            bk['client_status'],

            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Bus Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Reason For Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",
        'Approver1 Action',
        'Approver1 Name',
        'Approver1 Date',
        'Approver1 Time',
        'Approver2 Action',
        'Approver2 Name',
        'Approver2 Date',
        'Approver2 Time',
        'Approved Date',
        'Approved Time',
        'Approved By',
        'Canceled Date',
        'Canceled Time',
        'Canceled By',
        'Assigned Date',
        'Assigned Time',
        'Assigned By',
        'Passanger Name',
        'Pickup City',
        'Drop City',
        'Journey Date',
        'Journey Time',
        'Current Booking Status',
        'Bus Type Allocated',
        'PNR Number',
        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',
        'Is Auto Approved',
        'Bill ID',
        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by=''
        spoc_canceled_date=''
        spoc_status=''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']


                # Define the data for each cell in the row
        row = [

            bk['reference_no'],

            bk['assessment_code'],
            bk['assessment_city_id'],

            bk['reason_booking'],

            bk['zone_name'],

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            dateonly(bk['booking_datetime']),

            timeonly(bk['booking_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            passanger_list,

            bk['pickup_location'],

            bk['drop_location'],

            dateonly(bk['pickup_from_datetime']),

            timeonly(bk['pickup_from_datetime']),

            bk['cotrav_status'],

            bk['assign_bus_type_id'],

            bk['pnr_no'],

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            '',

            '',

            bk['client_status'],
            bk['cotrav_status'],

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "Pickup City",
        "Drop City",
        "Booking Date",
        "Booking Time",
        "Journey Date",
        "Journey Time",
        "SPOC Status",
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        "Zone",
        "Coach Type Allocated",
        "Quota Used",

        'No of seats',
        'Operator name',
        'Operator contact',
        'Train name',
        'Ticket no',
        'pnr no',
        'Assign bus type id',
        'Seat no',
        'Portal used',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status=''
        spoc_canceled_by=''
        spoc_canceled_date=''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            bk['zone_name'],
            bk['train_type_priority_1'],
            bk['seat_no'],


            bk['no_of_seats'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['train_name'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['seat_no'],
            bk['portal_used'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "From City",
        "To City",
        "Booking Date",
        "Booking Time",
        "Departure Date",
        "Departure Time",

        "Return Date",
        "Booking Status",

        "SPOC Status",
        "SPOC Cancel By",
        "SPOC Cancel Date",

        "Usage Type",
        "Trip Type",
        "Flight Type",
        "Seat Type",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        'First Flight Name',
        'First Flight No',
        'First Flight PNR No.',
        'First Flight From',
        'First Flight To',
        'First Flight Departure Date',
        'First Flight Departure time',
        'First Flight Arrival Date',
        'First Flight Arrival time',

        'Second Flight Name',
        'Second Flight No',
        'Second Flight PNR No.',
        'Second Flight From',
        'Second Flight To',
        'Second Flight Departure Date',
        'Second Flight Departure Time',
        'Second Flight Arrival Datetime',
        'Second Flight Arrival Time',

        'Third Flight Name',
        'Third Flight No',
        'Third Flight PNR No.',
        'Third Flight From',
        'Third Flight To',
        'Third Flight Departure Date',
        'Third Flight Departure Time',
        'Third Flight Arrival Date',
        'Third Flight Arrival Time',

        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        spoc_status= ''
        spoc_canceled_by=''
        spoc_canceled_date=''


        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            "Booking Remarks",
            bk['from_location'],
            bk['to_location'],

            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),

            "",
            bk['cotrav_status'],

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            flight_name1,
            flight_no1,
            pnr_no1,
            from_city1,
            to_city1,
            dateonly(departure_datetime1),
            timeonly(departure_datetime1),
            dateonly(arrival_datetime1),
            timeonly(arrival_datetime1),

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            dateonly(departure_datetime2),
            timeonly(departure_datetime2),
            dateonly(arrival_datetime2),
            timeonly(arrival_datetime2),

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            dateonly(departure_datetime3),
            timeonly(departure_datetime3),
            dateonly(arrival_datetime3),
            timeonly(arrival_datetime3),

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']



    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Bucket Request Date',
        'Bucket Request Time',
        'Bucket Approved Date',
        'Bucket Approved Time',
        'Preferred Hotel',
        'Assigned Hotel',
        'Assigned Hotel Address',
        'Hotel Contact',
        'Assign Date',
        'Assign Time',
        'TaxiVaxi Status',
        'Employees Name',
        'No. of Persons',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Higher Bucket Requested',
        'Reason for Higher Bucket',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',
        'Current Booking Status',
        'No. of Nights',
        'Room Type',
        'Room Occupancy',
        'Per Night Price',
        'Total Room Price',
        'Tax On Room Cancellation',
        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',
        'Is Auto Approved',
        'Bill ID',
        'Is TBA Booking',
        'Is Offline Booking',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")


        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            '',
            '',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            '',
            '',
            '',
            '',

            bk['preferred_hotel'],
            bk['assign_hotel_id'],
            bk['operator_name'],
            bk['operator_contact'],
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['status_cotrav'],
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),
            bk['status_cotrav'],
            '',
            bk['room_type_name'],
            bk['hotel_type_name'],
            bk['bucket_price_1'],
            bk['bucket_price_1'],
            '',
            bk['ticket_price'],

            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['sub_total'],
            bk['billing_entity_id'],
            bk['igst'],
            bk['cgst'],
            bk['sgst'],
            bk['management_fee_igst'],
            bk['management_fee_cgst'],
            bk['management_fee_sgst'],
            bk['management_fee_igst_rate'],
            bk['management_fee_cgst_rate'],
            bk['management_fee_sgst_rate'],

            '',
            '',
            '',
            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],


        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_billing_entities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Entitys']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=BillingEntitys.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Billing Entitys'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Billing City",
        "Entity Name",
        "Contact Person Name",
        "Contact Person Email",
        "Contact Person Phone No",
        "Address Line 1",
        "Address Line 2",
        "Address Line 3",
        "GST NO",
        "PAN No",
        "Is Active",
    ]

    row_num = 1



    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"
        # Define the data for each cell in the row
        row = [

            row_num-1,
            bk['corporate_name'],
            bk['billing_city'],
            bk['entity_name'],
            bk['contact_person_name'],
            bk['contact_person_email'],
            bk['contact_person_no'],
            bk['address_line_1'],
            bk['address_line_2'],
            bk['address_line_3'],
            bk['gst_id'],
            bk['pan_no'],
            is_deleted,
            
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_rates(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Corporate_Retes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyRates.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Rates'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City",
        "Taxi Type",
        "Tour Type",
        "KMS",
        "HOURS",
        "KM Rate",
        "HR Rate",
        "Base Rate",
        "Night Rate",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['tour_type'] == 2:
            tour_type = "Local"
        else:
            tour_type = "Outstation"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            bk['taxi_type'],
            tour_type,
            bk['kms'],
            bk['hours'],
            bk['km_rate'],
            bk['hour_rate'],
            bk['base_rate'],
            bk['night_rate'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_assessment_cities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "assessment_cities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Cities']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=AssessmentCities.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Assessment Cities'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_assessment_codes(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "assessment_codes"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Codes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=AssessmentCodes.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Assessment Codes'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Assessment Code",
        "Code Description",
        "From Date",
        "To Date",
        "Service From Date",
        "Service To Date",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_active'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['assessment_code'],
            bk['code_desc'],
            bk['from_date'],
            bk['to_date'],
            bk['service_from'],
            bk['service_to'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_groups(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Groups']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyGroups.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Groups'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Zone Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['zone_name'],

            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_subgroups(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Subgroups']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySubGroups.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company SubGroups'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgoup Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],

            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_admins(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Admins']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyAdmins.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Admins'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Admin Name",
        "Admin Email",
        "Admin Contact No ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['name'],
            bk['email'],
            bk['contact_no'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_spocs(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Spocs']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySpocs.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Spocs'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgroup Name",
        "Spoc Company ID",
        "Spoc Name",
        "Spoc Email",
        "Spoc Contact No ",
        "Login UserName ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['user_cid'],
            bk['user_name'],
            bk['email'],
            bk['user_contact'],
            bk['username'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_employees(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Employees']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyEmployees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Employees'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Spoc Name",
        "Core Employee ID",
        "Employee Company ID",
        "Employee Name",
        "Employee Email",
        "Employee Phone No",
        "Login UserName",
        "Age",
        "Gender",
        "ID Proof Type",
        "ID No",
        "Is CXO",
        "Designation",
        "Home City",
        "Home Address",
        "Assistant ID",
        "Date Of Birth",
        "Billing Entity",
        "Last Login",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_cxo'] == 0:
            is_cxo = "No"
        else:
            is_cxo = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['user_name'],
            bk['core_employee_id'],
            bk['core_employee_id'],
            bk['employee_cid'],
            bk['employee_name'],
            bk['employee_email'],
            bk['employee_contact'],
            bk['username'],
            bk['age'],
            bk['gender'],
            bk['id_proof_type'],
            bk['id_proof_no'],
            is_cxo,
            bk['designation'],
            bk['home_city'],
            bk['home_address'],
            bk['assistant_id'],
            bk['date_of_birth'],
            bk['billing_entity_id'],
            bk['last_login'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def taxi_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_taxi_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_taxi_billing.html", {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_taxi_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def taxi_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)

            vry_url = settings.API_BASE_URL + "admin_verify_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            print("current url")
            print(current_url)
            vry_url = settings.API_BASE_URL + "admin_revise_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def bus_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_bus_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_bus_billing.html", {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_bus_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def train_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_train_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_train_billing.html",
                          {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_train_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def flight_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_flight_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_flight_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_flight_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def hotel_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_hotel_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_hotel_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_hotel_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def visa_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "company_visa_services_bill"
        payload = {'visa_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Visa']
            return render(request, "Company/Admin/admin_visa_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_visa_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def frro_billing(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "company_frro_services_bill"
        payload = {'visa_type': id, 'user_id':request.user.id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Visa']
            return render(request, "Company/Admin/spoc_frro_billing.html",{'bookings': booking,'billing_type': id})
        else:
            return render(request, "Company/Admin/spoc_frro_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def bus_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def train_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def hotel_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def visa_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_visa_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_visa_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")

def frro_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_frro_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_frro_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def flight_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def accept_bill(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            current_url = request.POST.get('current_url', '')
            user_comment = request.POST.get('user_comment', '')

            url = ""
            operation_message = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "admin_accept_bill"
                operation_message="Bill Accepted successfully..!"

            if reject_id == '1':
                url = settings.API_BASE_URL + "admin_reject_bill"
                operation_message="Bill Rejected successfully..!"

            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, 'Failed to Accept Bill..!')
                return HttpResponseRedirect(current_url, {})
        else:
            return HttpResponseRedirect("/Corporate/Admin/bill/2")
    else:
        return HttpResponseRedirect("/login")

def get_all_generated_bills(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_all_generated_bills"
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = company['Bill']
        return render(request, 'Company/Admin/bills_geterated.html', {'bills': companies, 'bill_type':id})
    else:
        return HttpResponseRedirect("/login")


def corporate_bank_accounts(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        payload = {'corporate_id': request.user.corporate_id}
        url = settings.API_BASE_URL + "get_corporate_accounts"
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        pos = ''
        companies = ''
        if company['success'] == 1:
            pos = company['Accounts']
        else:
            pos = ''

        return render(request, 'Company/Admin/corporate_accounts.html', {'accounts': pos})
    else:
        return HttpResponseRedirect("/login")


def add_company_accounts(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            current_url = request.POST.get('current_url', '')
            corporate_id = request.POST.get('corporate_id', '')
            bank_name = request.POST.get('bank_name')
            bank_branch = request.POST.get('bank_branch', '')
            acoount_no = request.POST.get('acoount_no', '')
            acoount_holder_name = request.POST.get('acoount_holder_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')
            micr_code = request.POST.get('micr_code', '')

            account_id = request.POST.get('account_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'bank_name': bank_name, 'bank_branch': bank_branch, 'acoount_no': acoount_no, 'acoount_holder_name': acoount_holder_name,
                       'ifsc_code': ifsc_code,'micr_code':micr_code,  'account_id': account_id, 'is_delete': delete_id, }

            url = ""
            if account_id:
                url = settings.API_BASE_URL + "update_corporate_account"
                operation_message = "Company Account Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_corporate_account"
                    operation_message = "Company Account Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_corporate_account"
                operation_message = "Company Account Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def dashboard_search_admin_api_call(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        serveType = int(request.POST.get('serveType', ''))

        corp_id = request.user.corporate_id

        booking_type = int(request.POST.get('booking_type', ''))

        bookings_from_date = request.POST.get('bookings_from_date', '')

        bookings_to_date = request.POST.get('bookings_to_date', '')

        booking_id = ""
        pickup_location = ""
        pickup_date1 = ""
        pickup_date2 = ""
        pickup_date3 = ""
        pickup_date4 = ""
        spoc_id = ""
        operator_name = ""

        city = ""
        pnr_no = ""
        ass_code = ""

        checkin_date = ""
        voucher_no = ""
        hotel_name = ""

        search_serve_url = ""

        whereClause = "b.corporate_id = '" + str(corp_id) + "' "

        if serveType == 1:
            booking_id = request.POST.get('booking_id', '')
            pickup_location = request.POST.get('pickup_location', '')
            pickup_date1 = request.POST.get('pickup_date1', '')
            spoc_id = request.POST.get('spoc_id', '')
            operator_name = request.POST.get('operator_name', '')
            current_url = request.POST.get('current_url', '')

        if serveType == 2:
            booking_id = request.POST.get('booking_id', '')
            pickup_date2 = request.POST.get('pickup_date2', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')

        if serveType == 3:
            booking_id = request.POST.get('booking_id', '')
            pickup_date3 = request.POST.get('pickup_date3', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')

        if serveType == 4:
            booking_id = request.POST.get('booking_id', '')
            pickup_date4 = request.POST.get('pickup_date4', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/flight_bookings.html'
        if serveType == 5:
            booking_id = request.POST.get('booking_id', '')
            checkin_date = request.POST.get('checkin_date', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            ass_code = request.POST.get('ass_code', '')
            voucher_no = request.POST.get('voucher_no', '')
            hotel_name = request.POST.get('hotel_name', '')
            current_url = request.POST.get('current_url', '')

        if (bookings_from_date and bookings_to_date):
            bookings_from_date = bookings_from_date + ' 00:00:00'
            bookings_to_date = bookings_to_date + ' 00:00:00'

            bookings_from_date_object = datetime.strptime(bookings_from_date, '%d-%m-%Y %H:%M:%S')
            bookings_to_date_object = datetime.strptime(bookings_to_date, '%d-%m-%Y %H:%M:%S')

            bookings_from_date = bookings_from_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            bookings_to_date = bookings_to_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            print(bookings_from_date)

            print(bookings_to_date)

            if serveType == 1:
                whereClause = whereClause + "AND " + "b.booking_date BETWEEN CAST('" + bookings_from_date + "' AS DATE) AND CAST('" + bookings_to_date + "' AS DATE) "
            else:
                whereClause = whereClause + "AND " + "b.booking_datetime BETWEEN CAST('" + bookings_from_date + "' AS DATE) AND CAST('" + bookings_to_date + "' AS DATE) "

        if pickup_location:
            whereClause = whereClause + "AND " + "b.pickup_location LIKE '%" + pickup_location + "%' "

        if pickup_date1:
            pickup_date1 = pickup_date1 + ' 00:00:00'
            pickup_date1_object = datetime.strptime(pickup_date1, '%d-%m-%Y %H:%M:%S')
            pickup_date1 = pickup_date1_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_datetime) = CAST('" + pickup_date1 + "' AS DATE) "

        if pickup_date2:
            pickup_date2 = pickup_date2 + ' 00:00:00'
            pickup_date2_object = datetime.strptime(pickup_date2, '%d-%m-%Y %H:%M:%S')
            pickup_date2 = pickup_date2_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_from_datetime) = CAST('" + pickup_date2 + "' AS DATE) "

        if pickup_date3:
            pickup_date3 = pickup_date3 + ' 00:00:00'
            pickup_date3_object = datetime.strptime(pickup_date3, '%d-%m-%Y %H:%M:%S')
            pickup_date3 = pickup_date3_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_from_datetime) = CAST('" + pickup_date3 + "' AS DATE) "

        if pickup_date4:
            pickup_date4 = pickup_date4 + ' 00:00:00'
            pickup_date4_object = datetime.strptime(pickup_date4, '%d-%m-%Y %H:%M:%S')
            pickup_date4 = pickup_date4_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.departure_datetime) = CAST('" + pickup_date4 + "' AS DATE) "

        if spoc_id:
            whereClause = whereClause + "AND " + "b.spoc_id = '" + spoc_id + "' "

        if operator_name:
            whereClause = whereClause

        if city:
            whereClause = whereClause + "AND " + "b.pickup_location LIKE '%" + city + "%' "

        if pnr_no:
            whereClause = whereClause + "AND " + "b.pnr_no = '" + pnr_no + "' "

        if ass_code:
            whereClause = whereClause + "AND " + "b.assessment_code = '" + ass_code + "' "

        if checkin_date:
            checkin_date = checkin_date + ' 00:00:00'
            checkin_date_object = datetime.strptime(checkin_date, '%d-%m-%Y %H:%M:%S')
            checkin_date = checkin_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.checkin_datetime) = CAST('" + checkin_date + "' AS DATE) "

        if voucher_no:
            whereClause = whereClause + "AND " + "b.voucher_no = '" + voucher_no + "' "

        if hotel_name:
            whereClause = whereClause + "AND " + "ht.name LIKE '%" + hotel_name + "%' "

        if booking_id:
            whereClause = "b.reference_no = '" + booking_id + "' "

        search_serve_url = "Company/Admin/dashboard_search_admin_result.html"

        payload = {'whereClause': whereClause, 'serveType': serveType, 'booking_type': booking_type}
        url = settings.API_BASE_URL + "dashboard_search_admin_bookings"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Search Result..!")
            return render(request, search_serve_url, {'bookings': verify['Result'], 'serveType': serveType})

        else:
            messages.error(request, "Sory for error...!")
            return HttpResponse("error")
    else:
        return HttpResponseRedirect("/login")


def reports_invoice(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0

            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            date_type = request.POST.get('date_type', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')

            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            payload = {'corporate_id': int(corporate_id), 'service_type': service_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}

            print("payload")

            print(payload)

            url = settings.API_BASE_URL + "admin_report_invoice"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            #print(operator)
            if operator['success'] == 1:
                operator = operator['Reports']
            else:
                operator = {}

            return render(request, 'Company/Admin/reports_invoice.html',
                          {'Reports': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Company/Admin/reports_invoice.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/login")



def reports_client_billing(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0
            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            bill_status = request.POST.get('bill_status', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'bill_status': bill_status, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "admin_report_client_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}

            return render(request, 'Company/Admin/reports_client_billing.html',
                          {'bills': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Company/Admin/reports_client_billing.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/login")


def download_invoice_reports(request):

    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        if not corporate_id:
            corporate_id = 0
        service_type = int(request.POST.get('service_type', ''))
        date_type = request.POST.get('date_type', '')
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')

        service_text = ""

        if service_type == 1:

            service_text = "taxi"

        elif service_type == 2:

            service_text = "bus"

        elif service_type == 3:

            service_text = "train"

        elif service_type == 4:

            service_text = "flight"

        elif service_type == 5:

            service_text = "hotel"

        else:

            service_text = "all"



        payload = {'corporate_id': int(corporate_id), 'service_type': service_type,'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
        print(payload)
        url = settings.API_BASE_URL + "admin_report_invoice"
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Reports']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report_invoice_'+ service_text +'_'+ from_date +'_'+ to_date + '.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Invoice Reports'

    # Define the titles for columns

    columns = [
            "Sr.No",
            "reference_no",
            "assessment_code",
             "assessment_city_id",
            "pickup_location",
            "drop_location",
            "spoc_id",
            "user_name",
            "user_contact",
            "booking_datetime",
            "boarding_datetime",
            "boarding_point",
            "portal_used",
            "operator_name",
            "operator_contact",
            "ticket_no",
            "pnr_no",
            "assign_bus_type_id",
            "ticket_price",
            "management_fee",
            "tax_on_management_fee",
            "tax_on_management_fee_percentage",
            "sub_total",
            "vi_ticket_price",
            "vender_commission",

            "vender_commission",
            "invoice_status",
    ]



    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1

        if service_type == 1:
            bk['booking_datetime'] = bk['booking_date']
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['portal_used'] = ''
            bk['ticket_no'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''
            bk['ticket_price'] = ''

        if service_type == 3:
            bk['vi_ticket_price'] = ''
            bk['vender_commission'] = ''
            bk['invoice_status'] = ''

        if service_type == 4:
            bk['assessment_city_id'] = ''
            bk['pickup_location'] = bk['from_location']
            bk['drop_location'] = bk['to_location']
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['portal_used'] = ''
            bk['operator_name'] = ''
            bk['operator_contact'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''

        if service_type == 5:
            bk['pickup_location'] = ''
            bk['drop_location'] = ''
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['ticket_no'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['pickup_location'],
            bk['drop_location'],
            bk['spoc_id'],
            bk['user_name'],
            bk['user_contact'],
            bk['booking_datetime'],
            bk['boarding_datetime'],
            bk['boarding_point'],
            bk['portal_used'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['ticket_price'],
            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['tax_on_management_fee_percentage'],
            bk['sub_total'],
            bk['vi_ticket_price'],


            bk['vender_commission'],

            bk['vender_commission'],
            bk['invoice_status'],

        ]





        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_client_bill_reports(request):

    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        if not corporate_id:
            corporate_id = 0
        service_type = int(request.POST.get('service_type', ''))
        bill_status = int(request.POST.get('bill_status', ''))
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')

        service_text = ""

        if service_type == 1:

            service_text = "taxi"

        elif service_type == 2:

            service_text = "bus"

        elif service_type == 3:

            service_text = "train"

        elif service_type == 4:

            service_text = "flight"

        elif service_type == 5:

            service_text = "hotel"

        else:

            service_text = "all"


        if bill_status == 1 :

            bill_text = "Unpaid"

        elif bill_status == 2 :

            bill_text = "partial"

        elif bill_status == 3:

            bill_text = "paid"

        else:
            bill_text = "All"


        payload = {'corporate_id': int(corporate_id), 'service_type': service_type,'bill_status': bill_status, 'from_date': from_date, 'to_date': to_date}
        print(payload)
        url = settings.API_BASE_URL + "admin_report_client_bills"
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bill']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report_client_bills_'+ service_text +'_'+ from_date +'_'+ to_date + '_' + bill_text +'.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Client Bill Reports'

    # Define the titles for columns

    columns = [
        "Sr.No",
        "CorporatemName ",
        "Bill Number",
        "No Of Invoices",
        "Cotrav Billing Entity",
        "Client  Billing Entity ",
        "Billing Type",
        "TDS  Deducted  By Client",
        "System  Calculated TDS",
        "IGST",
        "CGST",
        "SGST",
        "Total Amount",
        "Is Paid",
        "Payment Status",
        "Total GST Paid",
        "Management Fee",
        "Outstanding Pending Payment",
        "Paid Total Amount",
        "Balance Total Amount",
        "Advance Payment",
        "Is Offline",
        "Reimbursement Voucher",
        "ID Taxable Amount",
        "Nontaxable Amount",
        "PO Id",
        "Bill Created Date",
        "Bill Final Date",
        "User Comment",
        "Cotrav Status",
        "Client Status",
    ]



    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_id'],
            bk['bill_number'],
            bk['no_of_invoices'],
            bk['cotrav_billing_entity'],
            bk['client_billing_entity'],
            bk['billing_type'],
            bk['tds_deducted_by_client'],
            bk['system_calculated_tds'],
            bk['igst'],
            bk['cgst'],

            bk['sgst'],
            bk['total_amount'],
            bk['is_paid'],

            bk['payment_status'],
            bk['total_gst_paid'],
            bk['management_fee'],
            bk['outstanding_pending_payment'],

            bk['paid_total_amount'],
            bk['balance_total_amount'],
            bk['advance_payment'],

            bk['is_offline'],
            bk['reimbursement_voucher_id'],
            bk['taxable_amount'],
            bk['nontaxable_amount'],
            bk['po_id'],
            bk['bill_created_date'],
            bk['bill_final_date'],
            bk['user_comment'],
            bk['cotrav_status'],
            bk['client_status'],

        ]



        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def add_booking_feedback(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_id = request.POST.get('user_id', '')
        booking_type = request.POST.get('booking_type', '')
        user_rating = request.POST.get('user_rating', '')

        user_actual_rating = ''
        if int(user_rating) >= 4:
            user_rating_text_good = request.POST.get('user_rating_text_good', '')
            if user_rating_text_good == 'other':
                user_actual_rating = request.POST.get('user_rating_text_good_txt', '')
            else:
                user_actual_rating = user_rating_text_good
        else:
            user_rating_text_bad = request.POST.get('user_rating_text_bad', '')
            if user_rating_text_bad == 'other':
                user_actual_rating = request.POST.get('user_rating_text_bad_txt', '')
            else:
                user_actual_rating = user_rating_text_bad

        url = settings.API_BASE_URL + "add_booking_feedback"
        payload = {'booking_id': booking_id, 'user_id': user_id, 'user_actual_rating': user_actual_rating, 'booking_type':booking_type, 'user_rating':user_rating}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            messages.error(request, company['message'])
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, company['message'])
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def visa_bokings(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if request.method == 'POST':
            pass
        else:
            url = settings.API_BASE_URL+"company_visa_services"
            payload = {'user_id': request.user.corporate_id}
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                corporates_data = company['Visa']
                return render(request,"Company/Admin/visa_requests.html",{'visa_services':corporates_data})
            else:
                return render(request,"Company/Admin/visa_requests.html",{'visa_services':{}})
    else:
        return HttpResponseRedirect("/login")


def view_visa_requests(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if request.method == 'POST':
            current_url = request.POST.get('current_url')
            request_type = request.POST.get('request_type')
            visa_request_type = request.POST.get('visa_request_type')
            corporate_id = request.POST.get('corporate_id')
            spoc_id = request.POST.get('spoc_id')
            group_id = request.POST.get('group_id')
            subgroup_id = request.POST.get('subgroup_id')
            country_id = request.POST.get('country_id')
            visa_type = request.POST.get('visa_type')
            visa_duration = request.POST.get('visa_duration')
            purpose_of_trip = request.POST.get('purpose_of_trip')
            if purpose_of_trip == 'Other':
                purpose_of_trip = request.POST.get('purpose_of_trip_txt')
            current_country_id = request.POST.get('current_country_id')
            current_state_id = request.POST.get('current_state_id')
            consulate_office_id = request.POST.get('consulate_office_id')
            no_of_employees = request.POST.get('no_of_employees')
            employee_email_1 = request.POST.get('employee_email_1')
            no_of_family_member = request.POST.get('no_of_family_member')
            application_form_link = request.POST.get('application_form_link')

            employee_name_1 = request.POST.getlist('employee_id')

            emp_no_of_document = request.POST.getlist('emp_no_of_document')
            emp_document = request.POST.getlist('emp_document')
            emp_document_txt = request.POST.getlist('emp_document_txt')

            family_members_name = request.POST.getlist('family_members_name')
            family_members_relationship = request.POST.getlist('family_members_relationship')
            empf_no_of_document = request.POST.getlist('empf_no_of_document')
            empf_document = request.POST.getlist('empf_document')
            empf_document_path = request.POST.getlist('empf_document_path')
            empf_document_txt = request.POST.getlist('empf_document_txt')

            employee_application_form = []
            employee_docs = []
            employeef_application_form = []
            employeef_docs = []
            print(request.POST)
            final_emp_no = int(no_of_employees) + 1
            final_no_of_family = int(no_of_family_member) + 1
            if request.FILES:
                if request_type == 'Family':
                    for i in range(1, final_no_of_family):
                        if request.FILES['empf_application_' + str(i)] != '':
                            file_up2 = request.FILES['empf_application_' + str(i)]
                            booking_email1 = upload_visa_doc_get_path(file_up2)
                            employeef_application_form.append(booking_email1)
                            val_i = i - 1
                            finl_emp_no_of_document = int(empf_no_of_document[val_i]) + 1
                            for ii in range(1, finl_emp_no_of_document):
                                if request.FILES['empf_document_path_' + str(i) + "_" + str(ii)] != '':
                                    file_up11 = request.FILES['empf_document_path_' + str(i) + "_" + str(ii)]
                                    employee_docs11 = upload_visa_doc_get_path(file_up11)
                                    employeef_docs.append(employee_docs11)
                else:
                    for i in range(1, final_emp_no):
                        if request.FILES['emp_application_form_' + str(i)] != '':
                            file_up = request.FILES['emp_application_form_' + str(i)]
                            booking_email1 = upload_visa_doc_get_path(file_up)
                            employee_application_form.append(booking_email1)
                            val_i = i - 1
                            finl_emp_no_of_document = int(emp_no_of_document[val_i]) + 1

                            for ii in range(1, finl_emp_no_of_document):
                                if request.FILES['emp_document_path_' + str(i) + "_" + str(ii)] != '':
                                    file_up11 = request.FILES['emp_document_path_' + str(i) + "_" + str(ii)]
                                    employee_docs11 = upload_visa_doc_get_path(file_up11)
                                    employee_docs.append(employee_docs11)

            payload = {'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'country_id': country_id, 'visa_type': visa_type,
                       'purpose_of_trip': purpose_of_trip, 'current_country_id': current_country_id,
                       'current_state_id': current_state_id, 'consulate_office_id': consulate_office_id,
                       'no_of_employees': no_of_employees, 'employee_ids': employee_name_1,
                       'employee_email_1': employee_email_1, 'request_type': request_type,
                       'visa_request_type': visa_request_type, 'visa_duration': visa_duration,
                       'no_of_family_member': no_of_family_member, 'application_form_link': application_form_link,
                       'emp_no_of_document': emp_no_of_document, 'emp_document': emp_document,
                       'emp_document_txt': emp_document_txt, 'family_members_name': family_members_name,
                       'family_members_relationship': family_members_relationship,
                       'empf_no_of_document': empf_no_of_document, 'empf_document': empf_document,
                       'empf_document_path': empf_document_path,
                       'employee_application_form': employee_application_form, 'employee_docs': employee_docs,
                       'employeef_application_form': employeef_application_form, 'employeef_docs': employeef_docs,
                       'empf_document_txt': empf_document_txt, 'is_email': 1, 'is_sms': 1}
            print(payload)
            url_cities1 = settings.API_BASE_URL + "add_visa_requests"
            company = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Visa Service Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Visa Service ..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Fails"})

        else:
            payload = {'some': 'data', 'spoc_id': request.user.id, 'visa_id':id}

            url_cities1 = settings.API_BASE_URL + "view_visa_request"
            taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
            visas = taxies11sad['Visa']

            url_cities1 = settings.API_BASE_URL + "get_countries"
            taxies11 = getDataFromAPI(login_type, access_token, url_cities1, payload)
            Country = taxies11['Country']

            url_cities111 = settings.API_BASE_URL + "get_states"
            taxies1ds1 = getDataFromAPI(login_type, access_token, url_cities111, payload)
            states = taxies1ds1['State']

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']

            url_emp = settings.API_BASE_URL + "spoc_employee"
            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employees']

            return render(request, "Company/Admin/view_visa_request.html", {'visas':visas,'countrys': Country, 'states': states, 'cities': cities, 'employees': employees})

    else:
        return HttpResponseRedirect("/login")


def view_frro_request(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        if request.method == 'POST':
            pass
        else:
            payload = {'frro_id': id}
            url_cities1 = settings.API_BASE_URL + "view_frro_request"
            taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(taxies11sad)
            if taxies11sad['success'] == 1:
                visas = taxies11sad['Frro']
                return render(request,"Company/Admin/view_frro_request.html",{'visas':visas})
            else:
                return render(request,"Company/Admin/view_frro_request.html",{'visas':{}})
    else:
        return HttpResponseRedirect("/login")


def get_all_frro_requests(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        if request.method == 'POST':
            pass
        else:
            url = settings.API_BASE_URL + "company_frro_request"
            payload = {'request_type': id, 'user_id': request.user.corporate_id}
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                corporates_data = company['Frros']
                return render(request, "Company/Admin/frro_services.html", {'visa_services': corporates_data, 'booking_type': id})
            else:
                return render(request, "Company/Admin/frro_services.html", {'visa_services': {}})
    else:
        return HttpResponseRedirect("/login")


def add_new_frro_request(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if request.method == 'POST':
            current_url = request.POST.get('current_url')
            request_type = request.POST.get('country_id')
            corporate_id = request.POST.get('corporate_id')
            billing_entity_id = request.POST.get('billing_entity_id')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            employee_id = request.POST.get('employee_id')


            city_id = request.POST.get('city_id')
            registered_email = request.POST.get('registered_email')
            reason_of_booking = request.POST.get('reason_of_booking')
            registered_username = request.POST.get('registered_username')
            registered_password = request.POST.get('registered_password')

            father_name = request.POST.get('father_name')
            mother_name = request.POST.get('mother_name')
            spouse_name = request.POST.get('spouse_name')
            height = request.POST.get('height')
            relegion = request.POST.get('relegion')
            outind_address = request.POST.get('outind_address')
            outind_city = request.POST.get('outind_city')
            outind_country = request.POST.get('outind_country')
            ind_address = request.POST.get('ind_address')
            ind_city = request.POST.get('ind_city')
            ind_country = request.POST.get('ind_country')
            embarkation_place = request.POST.get('embarkation_place')
            embarkation_city = request.POST.get('embarkation_city')
            embarkation_country = request.POST.get('embarkation_country')
            date_of_arrival = request.POST.get('date_of_arrival')
            disembarkation_place = request.POST.get('disembarkation_place')
            mode_of_journey = request.POST.get('mode_of_journey')
            trip_number = request.POST.get('trip_number')
            purpose_of_visit = request.POST.get('purpose_of_visit')
            is_military = request.POST.get('is_military')

            passport_no = request.POST.get('passport_no')
            passport_issue_date = request.POST.get('passport_issue_date')
            passport_expiry_date = request.POST.get('passport_expiry_date')
            visa_number = request.POST.get('visa_number')
            visa_expiry_date = request.POST.get('visa_expiry_date')
            current_location = request.POST.get('current_location')
            current_stay_address = request.POST.get('current_stay_address')
            company_contract_expiry_date = request.POST.get('company_contract_expiry_date')
            formc_expiry_date = request.POST.get('formc_expiry_date')
            lease_status = request.POST.get('lease_status')

            no_of_employees = 1
            is_email = request.POST.get('is_email')
            is_sms = request.POST.get('is_sms')

            employee_docs = {}
            doc_files = {}
            qty_docs = request.POST.getlist('doc_select[]')
            file_up11 = request.FILES.getlist('document[]')
            final_emp_no = int(no_of_employees) + 1
            # print(request.FILES['document[]'])
            # print(file_up11)
            for i, val in enumerate(qty_docs):
                if (0 <= i) and (i < len(file_up11)):
                    print(i)
                    print(file_up11[i])
                    employee_docs11 = upload_frro_doc_get_path(file_up11[i])
                    key = qty_docs[i]
                    employee_docs[key] = employee_docs11
                    doc_files[key] = employee_docs11
                else:
                    key = qty_docs[i]
                    doc_files[key] = ""

            payload = {'user_id': request.user.id, 'user_type': login_type, 'corporate_id': corporate_id,
                       'spoc_id': spoc_id,
                       'is_email': is_email, 'is_sms': is_sms, 'group_id': group_id, 'entity_id': billing_entity_id,
                       'no_of_employees': no_of_employees, 'subgroup_id': subgroup_id, 'request_type': request_type,
                       'employee_ids': employee_id, 'employee_docs': json.dumps(employee_docs),
                       'doc_files': json.dumps(doc_files), 'father_name': father_name, 'mother_name': mother_name,
                       'spouse_name': spouse_name, 'height': height, 'relegion': relegion,
                       'outind_address': outind_address, 'outind_city': outind_city, 'outind_country': outind_country,
                       'ind_address': ind_address, 'ind_city': ind_city, 'ind_country': ind_country,
                       'embarkation_place': embarkation_place, 'embarkation_city': embarkation_city,
                       'embarkation_country': embarkation_country, 'date_of_arrival': date_of_arrival,
                       'disembarkation_place': disembarkation_place, 'mode_of_journey': mode_of_journey,
                       'trip_number': trip_number, 'purpose_of_visit': purpose_of_visit, 'is_military': is_military,
                       'city_id': city_id, 'registered_email': registered_email,
                       'reason_of_booking': reason_of_booking, 'registered_username': registered_username,
                       'registered_password': registered_password, 'passport_no': passport_no,
                       'passport_issue_date': passport_issue_date, 'passport_expiry_date': passport_expiry_date,
                       'visa_number': visa_number, 'visa_expiry_date': visa_expiry_date,
                       'current_location': current_location, 'current_stay_address': current_stay_address,
                       'company_contract_expiry_date': company_contract_expiry_date,
                       'formc_expiry_date': formc_expiry_date, 'lease_status': lease_status}
            print(payload)
            url_cities1 = settings.API_BASE_URL + "add_frro_requests"
            company = getDataFromAPI(login_type, access_token, url_cities1, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, company['message'])
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect(current_url, {'message': "Operation Fails"})

        else:
            payload = {'some': 'data', 'spoc_id': request.user.id, 'corporate_id':request.user.corporate_id}

            url_cities1 = settings.API_BASE_URL + "get_frro_request_type"
            taxies11 = getDataFromAPI(login_type, access_token, url_cities1, payload)
            req_types = {}

            if taxies11['success'] == 1:
                req_types = taxies11['Types']
            else:
                req_types = {}

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities["Cities"]

            url_stat = settings.API_BASE_URL + "get_states"
            states = getDataFromAPI(login_type, access_token, url_stat, payload)
            states = states["State"]

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city1 = settings.API_BASE_URL + "get_assessment_city"
            citiess = getDataFromAPI(login_type, access_token, url_city1, payload)
            citiess = citiess['AssCity']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            return render(request, "Company/Admin/add_frro_request.html",
                          {'req_types': req_types, 'spocs': spocs, 'states': states, 'cities': cities, 'citiess':citiess, 'ass_code':ass_code, 'entities':entities})

    else:
        return HttpResponseRedirect("/login")


def frro_change_document(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        current_url = request.POST.get('current_url', '')
        document_id = request.POST.get('document_id', '')

        global booking_email
        booking_email = ''
        if request.FILES:
            file_up = request.FILES.get('documnet_file', False)
            if file_up:
                file_up = request.FILES['documnet_file']
                booking_email = upload_frro_doc_get_path(file_up)
            else:
                booking_email = None
        else:
            booking_email = None

        payload = {'document_id': document_id, 'document_file': booking_email}
        print(payload)
        url_cities1 = settings.API_BASE_URL + "frro_change_document"
        taxies11sad = getDataFromAPI(login_type, access_token, url_cities1, payload)
        return HttpResponseRedirect(current_url, {})

    else:
        return HttpResponseRedirect("/login")


def view_employee_frro_details(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        payload = {'employee_id': id}
        url = settings.API_BASE_URL + "view_employee"
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = company['Employee']

        url2 = settings.API_BASE_URL + "view_employee_documents"
        company1 = getDataFromAPI(login_type, access_token, url2, payload)
        companies1 = company1['Employee']

        return render(request, "Company/Admin/view_employee_document_details.html", {'employees': companies, 'documents':companies1})

    else:
            return HttpResponseRedirect("/login")



def calculate_age(born):
    print(born)
    dt_str = datetime.strptime(born, '%d-%m-%Y')
    today = date.today()
    return today.year - dt_str.year - ((today.month, today.day) < (dt_str.month, dt_str.day))


def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

def air_price_re():
    print("")
    """
    <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/">
 <soapenv:Header/>
 <soapenv:Body>
  <air:AirPriceReq xmlns:com="http://www.travelport.com/schema/common_v50_0" xmlns:air="http://www.travelport.com/schema/air_v50_0" TargetBranch="P7038885" CheckOBFees="All" AuthorizedBy="user" TraceId="P7038885">
   <com:BillingPointOfSaleInfo OriginApplication="UAPI"/>
   <air:AirItinerary>

<air:AirSegment Key="rIOS8N4R2BKAfoBmBAAAAA==" Group="0" Carrier="AI" FlightNumber="1126" Origin="HYD" Destination="DEL" DepartureTime="2020-07-29T22:10:00.000+05:30" ArrivalTime="2020-07-29T23:55:00.000+05:30" FlightTime="105" Distance="781" ETicketability="Yes" Equipment="32B" ChangeOfPlane="false" ParticipantLevel="Secure Sell" LinkAvailability="true" PolledAvailabilityOption="Cached status used. Polled avail exists" OptionalServicesIndicator="false" NumberOfStops="1" AvailabilitySource="P" AvailabilityDisplayType="Fare Shop/Optimal Shop" ProviderCode="1G" />
         

   </air:AirItinerary>
   <air:AirPricingModifiers FaresIndicator="AllFares">

   </air:AirPricingModifiers>
   <com:SearchPassenger Code="ADT" Key="1" Age="50"/>

   <air:AirPricingCommand/>
  </air:AirPriceReq>
 </soapenv:Body>
</soapenv:Envelope>
    """


def travelport_flight_booking_req(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':

            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')

            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            AirSegString = request.POST.get('AirSegString', '')

            AirPricingSolution = request.POST.get('AirPricingSolution', '')

            msg = ""

            booking = {}

            AirPricingSolutionString = ""

            #C = cookies.SimpleCookie()

            #PriceSolution = request.session['PriceSolution']

            #AirPricingSolutionString = request.session['PriceSolution']

            print('decoding starts ....................')

            #base64_message = AirPricingSolution
            #base64_bytes = base64_message.encode('ascii')
            #message_bytes = base64.b64decode(base64_bytes)
            #messagee = message_bytes.decode('ascii')
            print("encoded message is  *************")
            #print(messagee)

            AirPricingSolutionString = AirPricingSolution

            spoc_id = spoc_details[0]
            #group_id = spoc_details[1]
            #subgroup_id = spoc_details[2]
            group_id = 1
            subgroup_id = 1

            #now = datetime.now()

            now = datetime.now()  # current date and time


            usage_type = 1
            trip_type = 1
            seat_type = 1
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            booking_datetime = "20-09-2020 00:00:00"
            booking_datetime = now.strftime("%d-%m-%Y %H:%M:%S")
            #booking_datetime = now.strftime("%d-%m-%Y %H:%M:%S")
            departure_date2 = request.POST.get('departure_date', '')
            departure_date = "29-09-2020"
            preferred_flight = "any"
            assessment_code = "3"
            assessment_city_id = "1"
            entity_id = "1"
            reason_booking = "test"
            no_of_seats = int(request.POST.get('no_of_seats', ''))
            #no_of_seats = 2

            print("Printing Departure date 2 *****************")
            print(departure_date2)

            print("printing no of seats are ***********************")
            print(no_of_seats)

            flight_class_is_international = 0
            emp_info_international = []
            emp_data = {}
            emp_add_str = """"""
            #AirPricingSolutionString = """"""

            #emp_phone_key = "RyYXZlbGVyMQQm9va2luZ1=="
            emp_phone_no = "111111"
            phone_country_code = "91"
            phone_location = "DEL"
            phone_extention = "22"
            phone_area_code = "222"
            phone_type = "Home"
            phone_text = "Abc-Xy"
            emp_title = ""
            emp_fname = ""
            emp_lname = ""
            emp_dob = ""
            emp_email = ""

            emp_add_str2 = """"""
            passengerTypeStr = """"""

            for i in range(1, no_of_seats + 1 ):
                print("printing employee dattaaaa****************************")
                emp_data['emp_title'] = request.POST.get('employee_title_' + str(i), '')
                emp_data['emp_fname'] = request.POST.get('employee_fname_' + str(i), '')
                emp_data['emp_lname'] = request.POST.get('employee_lname_' + str(i), '')
                emp_data['emp_dob'] = request.POST.get('employee_edob_' + str(i), '')
                emp_data['emp_email'] = request.POST.get('employee_email_' + str(i), '')

                emp_info_international.append(emp_data)
                emp_title = emp_data['emp_title']
                emp_fname = emp_data['emp_fname']
                emp_lname = emp_data['emp_lname']
                emp_dob = emp_data['emp_dob']
                emp_email = emp_data['emp_email']

                str_res = ''.join(random.choices(string.ascii_uppercase +
                                             string.digits, k=22))

                emp_key = ''.join(str(str_res) + "==" )

                str_ph_res = ''.join(random.choices(string.ascii_uppercase +
                                                 string.digits, k=22))

                emp_phone_key = ''.join(str(str_ph_res) + "==" )


                passengerTypeStr = passengerTypeStr + """<air:PassengerType Code="ADT" BookingTravelerRef=""" + '"' + emp_key  +'"' + """ />"""


                emp_add_str = """<com:BookingTraveler Key=""" + '"' + str(emp_key) + '"' + """ TravelerType="ADT" xmlns:com="http://www.travelport.com/schema/common_v50_0">""" + \
                                """<com:BookingTravelerName Prefix=""" + '"' + str(emp_title) + '"'+ \
                                """ First=""" + '"' + str(emp_fname) +'"' + """ Last=""" + '"' + str(emp_lname) + '"' + """ />""" + \
                                """<com:PhoneNumber Key=""" + '"' + str(emp_phone_key) + '"' + """ CountryCode=""" + '"' + str(phone_country_code) + '"' + \
                                """ Location=""" + '"' + str(phone_location) + '"' + """ Number=""" + '"'+ str(emp_phone_no) + '"' + \
                                """ Extension=""" + '"' + str(phone_extention) + '"' + """ AreaCode=""" + '"' + str(phone_area_code) + '"' + \
                                """ Type=""" + '"' + str(phone_type) + '"' + """ Text=""" + '"' + str(phone_text) + '"' + """ /> """ + \
                                """<com:Email Type=""" + '"' + str(phone_type) + '"' + """ EmailID=""" + '"' + str(emp_email) + '"' + """/>""" + \
                                """<com:SSR Key=""" + '"' + str(i) + '"' + """ Type="DOCS" Status="HK" Carrier="SA" FreeText=""" + '"' + """P/CA/F9850356/GB/04JAN80/M/01JAN14/LINDELOEV/CARSTENGJELLERUPMr""" + '"' + """ /> """ + \
                                """<com:Address>""" +\
                                  """<com:AddressName>""" + str(emp_fname) + """  """ + str(emp_lname)   +"""</com:AddressName>""" + \
                                  """<com:Street>""" + """123 M G Road""" + """</com:Street>""" + \
                                  """<com:Street>""" + """Defense Colony""" + """</com:Street>""" + \
                                  """<com:City>""" + """Delhi""" + """</com:City>""" + \
                                  """<com:State>""" + """Delhi""" + """</com:State>""" + \
                                  """<com:PostalCode>""" + """110024 """ + """</com:PostalCode>""" + \
                                  """<com:Country>""" + """IN""" + """</com:Country>""" + \
                                """</com:Address>""" + \
                                """</com:BookingTraveler>"""

                print(emp_add_str)

                emp_add_str2 = emp_add_str2 + emp_add_str

            try:
                #new_air_pricing_resp = AirPricingSolutionString

                print("Print at flag 1 ***************")

                print(AirPricingSolutionString)



                start11 = 0
                count11 = 0

                new_air_pricing_resp = AirPricingSolutionString

                newPassengerTypeStr = passengerTypeStr

                strt = 0

                if (AirPricingSolutionString.find('<air:PassengerType', strt) != -1) and no_of_seats > 1:
                    result1 = AirPricingSolutionString.find('<air:PassengerType', strt)

                    result2 = AirPricingSolutionString.find('/>', result1)

                    psg_tag = AirPricingSolutionString[result1:result2 + 2]

                    AirPricingSolutionString = AirPricingSolutionString.replace(psg_tag, newPassengerTypeStr)


                new_air_pricing_resp = AirPricingSolutionString

                print("Print at flag 2 *****************")
                print(new_air_pricing_resp)


                while (new_air_pricing_resp.find('<air:PassengerType', start11) != -1) and no_of_seats > 1:

                    result1 = new_air_pricing_resp.find('<air:PassengerType', start11)

                    result2 = new_air_pricing_resp.find('/>', result1)

                    start11 = result2 + 2

                    count11 = count11 + 1

                    new_air_pricing_resp6 = new_air_pricing_resp[result1:result2 + 2]

                    cnt = new_air_pricing_resp.count(new_air_pricing_resp6)

                    if cnt > 1 :
                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp6, '', cnt-1)


                if no_of_seats > 1 :
                    AirPricingSolutionString = new_air_pricing_resp



                print("Print at flag 3 *****************")
                print(new_air_pricing_resp)


                print("printing AirPricingSolutionString for passanger counting ****************************")
                print(AirPricingSolutionString)


                payload = """
                <soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
                  <soap:Body>
                    <univ:AirCreateReservationReq RetainReservation="Both" RetrieveProviderReservationDetails="true" TraceId="3BCE9B92" TargetBranch="P7009927" AuthorizedBy="UAPITESTING" Version="0" xmlns:univ="http://www.travelport.com/schema/universal_v50_0" xmlns:air="http://www.travelport.com/schema/air_v50_0">
                      <com:BillingPointOfSaleInfo OriginApplication="UAPI" xmlns:com="http://www.travelport.com/schema/common_v50_0" />""" + emp_add_str2 + \
                      """<GeneralRemark UseProviderNativeMode="true" TypeInGds="Basic" xmlns="http://www.travelport.com/schema/common_v50_0">
                        <RemarkData>Booking 1</RemarkData>
                      </GeneralRemark>
                      <GeneralRemark UseProviderNativeMode="true" TypeInGds="Basic" xmlns="http://www.travelport.com/schema/common_v50_0">
                        <RemarkData>Re- Booking 1</RemarkData>
                      </GeneralRemark>
                      <com:ContinuityCheckOverride Key="1T" xmlns:com="http://www.travelport.com/schema/common_v50_0">true</com:ContinuityCheckOverride>
                      <com:FormOfPayment xmlns:com="http://www.travelport.com/schema/common_v50_0" Type="Cash" />""" + \
                        AirPricingSolutionString + \
                      """<com:ActionStatus ProviderCode="1G" TicketDate="T*" Type="ACTIVE" QueueCategory="01" xmlns:com="http://www.travelport.com/schema/common_v50_0" />
                    </univ:AirCreateReservationReq>
                  </soap:Body>
                </soap:Envelope>
                """
                header = {
                    "Content-Type": "text/xml:charset=utf-8",
                    "Accept": "gzip,deflate",
                    "Connection": "Keep-Alive",
                    "Authorization": "Basic %s" % CREDENTIALS_enc64,
                    "Content-Length": str(len(payload))
                }

                url = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/AirService"

                print("Payload is ............................")
                print(payload)
                print("Payload end ............................")
                response = requests.post(url, data=payload, headers=header)
                fligh_response = response.text

                print('Flight Response *****************************')
                #print(fligh_response)

                handler = FlightBookingHandler()

                xml.sax.parseString(fligh_response, handler)

                var1 = handler.getFinalData()

                var2 = json.dumps(var1)

                var3 = json.loads(var2)

                print("printing dataaaa")

                print(type(var1))

                print(type(var2))

                print(type(var3))

                print(var3)

                ##############################################################

                if "error_flag" in var3:
                    print("printing error flag")
                    print(var3["error_flag"])
                    msg = "Error in flight booking"

                    messages.error(request, str(msg))

                    if "faultcode" in var3["error_flag"] and var3["error_flag"] == "400" :
                        msg = "Request data not found - Request ignored"
                        messages.error(request, str(msg))

                    if "faultstring" in var3:
                        msg2 = var3["faultstring"]
                        messages.error(request, str(msg2))


                else:

                ###################################################

                    if "LocatorCode" in var3:

                        print("printing Booking Details to assign *************************")
                        print(var3)
                        locator_code = var3["LocatorCode"]


                        payload2 = """<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"><soapenv:Body>""" + \
                                          """<univ:UniversalRecordRetrieveReq TargetBranch="P7009927" xmlns:univ="http://www.travelport.com/schema/universal_v50_0" xmlns:com="http://www.travelport.com/schema/common_v50_0">""" + \
                                             """<com:BillingPointOfSaleInfo OriginApplication="UAPI"/>""" + \
                                             """<univ:UniversalRecordLocatorCode>""" + locator_code  + """</univ:UniversalRecordLocatorCode>""" + \
                                          """</univ:UniversalRecordRetrieveReq></soapenv:Body></soapenv:Envelope>"""



                        url2 = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/UniversalRecordService"

                        header2 = {
                            "Content-Type": "text/xml:charset=utf-8",
                            "Accept": "gzip,deflate",
                            "Connection": "Keep-Alive",
                            "Authorization": "Basic %s" % CREDENTIALS_enc64,
                            "Content-Length": str(len(payload2))
                        }

                        print("Payload is ............................")
                        print(payload2)
                        print("Payload end ............................")
                        response = requests.post(url2, data=payload2, headers=header2)
                        universal_fligh_record_retrive_response = response.text

                        print('Printing universal flight record *****************************')

                        print(universal_fligh_record_retrive_response)

                        no_of_segments = universal_fligh_record_retrive_response.count('</air:AirSegment>')

                        print("no of segments...........**********")

                        print(no_of_segments)

                        handler = UniversalRecordRetrieveRsp()

                        xml.sax.parseString(universal_fligh_record_retrive_response, handler)

                        retrive_air_var1 = handler.getFinalData()

                        retrive_air_var2 = json.dumps(retrive_air_var1)

                        retrive_air_var3 = json.loads(retrive_air_var2)

                        AirReservationLocatorCode = retrive_air_var3["AirReservationLocatorCode"] if "AirReservationLocatorCode" in retrive_air_var3 else ""
                        ProviderLocatorCode = retrive_air_var3["ProviderLocatorCode"] if "ProviderLocatorCode" in retrive_air_var3 else ""
                        base_price = retrive_air_var3["BasePrice"] if "BasePrice" in retrive_air_var3 else "INR0"

                        BasePrice1 = base_price.replace("INR","")

                        BasePrice = int(BasePrice1)

                        if entity_id:
                            pass
                        else:
                            entity_id = 0

                        employees = []
                        no_of_emp = int(no_of_seats) + 1
                        for i in range(1, no_of_emp):
                            employees.append(request.POST.get('employee_id_' + str(i), ''))
                            print(employees)

                        payload = {'user_id': user_id, 'user_type': login_type, 'corporate_id': corporate_id,
                                   'spoc_id': spoc_id, 'group_id': group_id,
                                   'subgroup_id': subgroup_id, 'usage_type': usage_type, 'trip_type': trip_type,
                                   'seat_type': seat_type, 'from_city': from_city, 'to_city': to_city,
                                   'booking_datetime': booking_datetime, 'departure_datetime': departure_date,
                                   'preferred_flight': preferred_flight, 'assessment_code': assessment_code,
                                   'reason_booking': reason_booking, 'no_of_seats': no_of_seats, 'employees': employees,
                                   'billing_entity_id': entity_id,'is_self_booking': 1 ,
                                   'is_sms': 1, 'is_email': 1, 'assessment_city_id': assessment_city_id}
                        print(payload)

                        url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
                        booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
                        print(booking)

                        ##################

                        if booking['success'] == 1:
                            last_booking_id = booking['last_booking_id']
                            booking_reference_no = booking['booking_reference_no']
                            last_booking_id = last_booking_id

                            ticket_number = []
                            pnr_no = []
                            flight_no = []
                            flight_name = []
                            arrival_time = []
                            departure_time = []
                            flight_to = []
                            flight_from = []
                            #is_return_flight = []
                            no_of_stops = "1"
                            seat_type = ""
                            flight_type = "onword"
                            fare_type = ""
                            meal_is_include = ""
                            no_of_passanger = str(no_of_seats)
                            employee_booking_id = ""
                            ticket_price = BasePrice/no_of_seats
                            sub_total = BasePrice * no_of_seats
                            vendor_booking = ""
                            is_return_flight = []

                            if "AirSegments" in retrive_air_var3["AirReservation"]:

                                air_pnr_no = retrive_air_var3["AirReservation"]["LocatorCode"]

                                for seg in retrive_air_var3["AirReservation"]["AirSegments"]:
                                    fl_name = seg["Carrier"] + "-" + seg["FlightNumber"]
                                    ar_t = seg["ArrivalTime"]
                                    dp_t = seg["DepartureTime"]
                                    #ArrivalTime = dateutil.parser.parse(ar_t)
                                    #DepartureTime = dateutil.parser.parse(dp_t)
                                    #datetime.datetime.strptime(departure_time[x], '%d-%m-%Y %H:%M')

                                    print(ar_t)

                                    print(dp_t)

                                    #ar_t_obj = datetime.strptime(ar_t, "%Y-%m-%dT%H:%M:%S.%f%z")

                                    #dp_t_obj = datetime.strptime(dp_t, "%Y-%m-%dT%H:%M:%S.%f%z")

                                    ar_t_obj = dateutil.parser.isoparse(str(ar_t))

                                    dp_t_obj = dateutil.parser.isoparse(str(dp_t))

                                    ArrivalTime = ar_t_obj.strftime("%d-%m-%Y %H:%M")

                                    DepartureTime = dp_t_obj.strftime("%d-%m-%Y %H:%M")

                                    #ArrivalTime = datetime.strptime(ar_t_obj, "%d-%m-%Y %H:%M") # 2020-10-22T23:50:00.000+05:30
                                    #DepartureTime = datetime.strptime(dp_t_obj, "%d-%m-%Y %H:%M") #2020-10-29T21:15:00.000+05:30

                                    print(ArrivalTime)
                                    print(DepartureTime)

                                    ticket_number.append("1111")
                                    pnr_no.append(air_pnr_no)
                                    flight_no.append(seg["FlightNumber"])
                                    flight_name.append(fl_name)
                                    arrival_time.append(ArrivalTime)
                                    departure_time.append(DepartureTime)
                                    flight_to.append(seg["Destination"])
                                    flight_from.append(seg["Origin"])
                                    is_return_flight.append(seg["Group"])

                                    print("is_return_flight ***************")
                                    print(is_return_flight)

                                    print("Printing segment group ***************")
                                    print(seg["Group"])


                            if no_of_segments > 1 :
                                no_seg = no_of_segments - 1
                                no_of_stops = str(no_seg)
                            else:
                                no_of_stops = "0"


                            url_assign = settings.API_BASE_URL + "assign_flight_booking"
                            payload11 = {'ticket_no': ticket_number, 'pnr_no': pnr_no, 'portal_used': "",
                                         'booking_id': last_booking_id, 'user_id': user_id, 'user_type': login_type,
                                         'flight_no': flight_no,
                                         'flight_name': flight_name, 'arrival_time': arrival_time,
                                         'departure_time': departure_time, 'flight_to': flight_to,
                                         'flight_from': flight_from,
                                         'no_of_stops': no_of_stops, 'seat_type': seat_type, 'flight_type': flight_type,
                                         'trip_type': trip_type, 'fare_type': fare_type,
                                         'meal_is_include': meal_is_include,
                                         'no_of_passanger': no_of_passanger, 'employee_booking_id': employee_booking_id,
                                         'ticket_price': ticket_price, 'management_fee': '100',
                                         'tax_mng_amt': '18', 'tax_on_management_fee': '18',
                                         'tax_on_management_fee_percentage': '18',
                                         'sub_total': sub_total,
                                         'management_fee_igst': 18, 'management_fee_cgst': 0,
                                         'management_fee_sgst': 0,
                                         'management_fee_igst_rate': int(ticket_price) * 0.18,
                                         'management_fee_cgst_rate': 0,
                                         'management_fee_sgst_rate': 0, 'cgst': 0, 'sgst': 0,
                                         'igst': int(sub_total) * 0.18,
                                         'oper_ticket_price': ticket_price,
                                         'oper_commission': "",
                                         'oper_commission_type': "",
                                         'oper_cotrav_billing_entity': "1",
                                         'cotrav_billing_entity': '1',
                                         'oper_cgst': 0, 'oper_sgst': 0, 'oper_igst': 18,
                                         'client_ticket_path': '', 'client_ticket': '1',
                                         'vender_ticket': '',
                                         'vender_ticket_path': '', 'is_client_sms': '1',
                                         'is_client_email': '1',
                                         'igst_amount': int(ticket_price) * 0.18, 'cgst_amount': 0, 'sgst_amount': 0,
                                         'operator_id': '1', 'vendor_booking_id': vendor_booking,
                                         'is_return_flight': is_return_flight,
                                         }
                            print("payrol  .....")
                            print(payload11)
                            company11 = getDataFromAPI(login_type, access_token, url_assign, payload11)
                            print(company11)

                            if company11['success'] == 1:
                                messages.success(request, 'Flight Booking Added Successfully..!')

                                add_code_payload = {'book_id': last_booking_id , 'book_ref': booking_reference_no , "uni_record_locator" : locator_code ,"provider_locator_code": ProviderLocatorCode , "reservation_locator_code": AirReservationLocatorCode }
                                url_add_code = settings.API_BASE_URL + "add_travelport_locator_code"

                                data = getDataFromAPI(login_type, access_token, url_add_code, add_code_payload)

                                if data["success"] == 1 :
                                    messages.success(request, 'Locator Codes are saved..!')
                                else:

                                    messages.success(request, 'Locator Codes are saved..!')

                            else:

                                messages.success(request, 'Flight is booked but error to assign details ..!')


                        ###################


                        msg = "Flight Booking is successful , Your PNR No is " + str(var3["LocatorCode"])
                        messages.success(request, str(msg))

                    else:

                        msg = "Error in Flight Booking"
                        messages.error(request, str(msg))

                ###################################################

                if "LocatorCode" in var3 and booking['success'] == 1:

                    if "message" in booking:
                        messages.success(request, str(booking['message']))

                    return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2",{'message': msg})
                    #return render(request, 'Company/Admin/api_call.html', {"response": var3, "UID2": 1 , 'message': "Operation Successfully" })
                else:
                    messages.error(request, 'Failed to Add Flight Booking..!')

                    return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2",{'message': msg})
                    #return render(request, 'Company/Admin/api_call.html', {"response": var3, "UID2": 1 , 'message': "Operation Successfully" })


                ########################################################################


            except Exception as e:
                messages.error(request, e)
                print(traceback.print_exc())
                return redirect('travelport_flight_booking_req')
        else:

            if 'admin_login_type' in request.session:
                login_type = request.session['admin_login_type']
                access_token = request.session['admin_access_token']
                payload = {'corporate_id': id}


            return render(request, 'Company/Admin/travelport_flight_search.html',{} )

    else:
        return HttpResponseRedirect("/login")


def travelport_cancle_booking_req(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':

            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            booking_id = request.POST.get('book_id', '')

            emp_id_list = request.POST.get('emp_id_list', '')

            print("printing employee losting **************")

            print(emp_id_list)

            #print(json.dumps(emp_id_list))

            D2 = eval(emp_id_list)

            employee_ar = json.dumps(D2)

            emp_dataa = json.loads(employee_ar)

            print(type(emp_dataa))

            new_emp_dict = []

            #print(type(employee_ar))

            emp_count = 0

            for emp in emp_dataa:
                emp_count = emp_count + 1
                new_emp_dict.append(emp['id'])


            print("printing new employee id list *****************")
            print(new_emp_dict)

            no_of_passenger = emp_count

            msg = ""

            try:
                add_code_payload = {'book_id': booking_id}
                url_add_code = settings.API_BASE_URL + "get_travelport_locator_code"

                data = getDataFromAPI(login_type, access_token, url_add_code, add_code_payload)
                flight_data = data['Codes']
                print(flight_data)

                print("Provider Locator Code is *******")
                print(flight_data[0]["provider_locator_code"])

                provider_locator = flight_data[0]["provider_locator_code"]

                univerasal_record_locator = flight_data[0]["univerasal_record_locator"]

                payload = """
                              <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:univ="http://www.travelport.com/schema/universal_v50_0" xmlns:com="http://www.travelport.com/schema/common_v50_0">
   <soapenv:Header/>
   <soapenv:Body>
      <univ:UniversalRecordCancelReq TargetBranch="P7009927" AuthorizedBy="SUSIL" UniversalRecordLocatorCode=""" + '"' + univerasal_record_locator + '"' + """ Version="1">"""  + \
         """<com:BillingPointOfSaleInfo OriginApplication="UAPI"/>
      </univ:UniversalRecordCancelReq>
   </soapenv:Body>
</soapenv:Envelope>
              """
                header = {
                    "Content-Type": "text/xml:charset=utf-8",
                    "Accept": "gzip,deflate",
                    "Connection": "Keep-Alive",
                    "Authorization": "Basic %s" % CREDENTIALS_enc64,
                    "Content-Length": str(len(payload))
                }

                url = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/UniversalRecordService"

                print("Payload is ............................")
                print(payload)
                print("Payload end ............................")
                response = requests.post(url, data=payload, headers=header)
                fligh_response = response.text

                print('Flight Response *****************************')
                print(fligh_response)

                #handler = UniversalRecordRetrieveRsp()

                handler = UniversalRecordCancelRsp()

                xml.sax.parseString(fligh_response, handler)

                var1 = handler.getFinalData()

                var2 = json.dumps(var1)

                var3 = json.loads(var2)

                print("printing dataaaa")

                print(type(var1))

                print(type(var2))

                print(type(var3))

                print(var3)


                if "error_flag" in var3:

                    print("printing error flag")
                    print(var3["error_flag"])
                    msg = "Error in flight booking"

                    messages.error(request, str(msg))

                    if "faultcode" in var3["error_flag"] and var3["error_flag"] == "400" :
                        msg = "Request data not found - Request ignored"
                        messages.error(request, str(msg))

                    if "faultstring" in var3:
                        msg2 = var3["faultstring"]
                        messages.error(request, str(msg2))

                else:

                    ##############################################################
                    tax_on_management_fee = 0
                    refund_amount = 0
                    cancel_comment = 0
                    management_fee_igst = 0
                    management_fee_cgst = 0
                    management_fee_sgst = 0
                    management_fee_igst_rate = 0
                    management_fee_cgst_rate = 0
                    management_fee_sgst_rate = 0
                    cgst = 0
                    sgst = 0
                    igst = 0
                    igst_amount = 0
                    cgst_amount = 0
                    sgst_amount = 0

                    user_id = request.user.id
                    no_of_passenger = emp_count
                    employees = new_emp_dict
                    refund_amount = "123"
                    cancel_comment = "12"
                    igst_rate = request.POST.get('igst_rate', '')

                    if igst_rate:
                        igst_rate = int(igst_rate)
                    else:
                        igst_rate = 0

                    ticket_price = request.POST.get('ticket_price', '')
                    if ticket_price:
                        pass
                    else:
                        ticket_price = 0
                    old_ticket_price = int(ticket_price)
                    ticket_price = int(ticket_price) - int(refund_amount)
                    management_fee = "50"
                    tax_mng_amt = ticket_price * 0.18
                    tax_on_management_fee = int(management_fee) * 0.18
                    tax_on_management_fee_percentage = 18
                    sub_total = ticket_price + int(
                        management_fee) + tax_mng_amt + tax_on_management_fee + tax_on_management_fee_percentage

                    if int(igst_rate) > 0:
                        management_fee_igst = 18
                        management_fee_cgst = 0
                        management_fee_sgst = 0
                        management_fee_igst_rate = 18
                        management_fee_cgst_rate = 0
                        management_fee_sgst_rate = 0
                        cgst = 18
                        sgst = 0
                        igst = 0
                        igst_amount = sub_total * 0.18
                        cgst_amount = 0
                        sgst_amount = 0
                    else:
                        management_fee_igst = 0
                        management_fee_cgst = 9
                        management_fee_sgst = 9
                        management_fee_igst_rate = 0
                        management_fee_cgst_rate = 9
                        management_fee_sgst_rate = 9
                        cgst = 0
                        sgst = 9
                        igst = 9
                        igst_amount = 0
                        cgst_amount = sub_total * 0.9
                        sgst_amount = sub_total * 0.9

                    url = settings.API_BASE_URL + "cancel_flight_booking_passengers"
                    payload = {'booking_id': booking_id, 'user_id': user_id, 'user_type': login_type,
                               'no_of_passenger': no_of_passenger,
                               'ticket_price': ticket_price, 'management_fee': management_fee, 'tax_mng_amt': tax_mng_amt,
                               'tax_on_management_fee': tax_on_management_fee, 'refund_amount': refund_amount,
                               'cancel_comment': cancel_comment,
                               'tax_on_management_fee_percentage': tax_on_management_fee_percentage, 'sub_total': sub_total,
                               'management_fee_igst': management_fee_igst, 'management_fee_cgst': management_fee_cgst,
                               'management_fee_sgst': management_fee_sgst,
                               'management_fee_igst_rate': management_fee_igst_rate,
                               'management_fee_cgst_rate': management_fee_cgst_rate, 'old_ticket_price': old_ticket_price,
                               'management_fee_sgst_rate': management_fee_sgst_rate, 'cgst': cgst, 'sgst': sgst,
                               'igst': igst,
                               'igst_amount': igst_amount, 'cgst_amount': cgst_amount, 'sgst_amount': sgst_amount,
                               'employees': employees
                               }
                    print(payload)
                    company = getDataFromAPI(login_type, access_token, url, payload)
                    print(company)
                    if company['success'] == 1:
                        messages.success(request, 'Employee Canceled Successfully')
                        # return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})


                    else:
                        messages.error(request, 'Employee Cancel Failed')
                        # return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
                        #return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2", {'message': msg})

                    #######################################################################

                    return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2", {'message': msg})

            except Exception as e:
                messages.error(request, e)
                print(traceback.print_exc())
                return redirect('travelport_flight_search')

        else:

            if 'admin_login_type' in request.session:
                login_type = request.session['admin_login_type']
                access_token = request.session['admin_access_token']
                payload = {'corporate_id': id}

            return render(request, 'Company/Admin/travelport_flight_search.html', {})

    else:

        return HttpResponseRedirect("/login")


def travelport_flight_details(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':

            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            Keyy = request.POST.getlist('Key')
            Group = request.POST.getlist('Group')
            Carrier = request.POST.getlist('Carrier')
            FlightNumber = request.POST.getlist('FlightNumber')
            Origin = request.POST.getlist('Origin')
            Destination = request.POST.getlist('Destination')
            DepartureTime = request.POST.getlist('DepartureTime')
            ArrivalTime = request.POST.getlist('ArrivalTime')
            FlightTime = request.POST.getlist('FlightTime')
            Distance = request.POST.getlist('Distance')

            no_of_seats = request.POST.get('no_of_seats' , '')
            departure_datetime = request.POST.get('departure_datetime', '')
            return_datetime = request.POST.get('return_datetime', '')
            fl_Origin = request.POST.get('Origin')
            fl_Destination = request.POST.get('Destination')

            new_air_pricing_resp = ""

            fl_class = ""

            OptionalServicesIndicator = request.POST.get('OptionalServicesIndicator', '')

            booking_datas = {"from_city": fl_Origin , "to_city": fl_Destination , "departure_datetime": departure_datetime ,"return_datetime": return_datetime, "fl_class": fl_class, "no_of_seats": no_of_seats}


            print("booking dataaaa********")
            print(booking_datas)

            print(Keyy)
            print(Group)
            print(Carrier)
            print(FlightNumber)
            print(Origin)
            print(Destination)
            print(DepartureTime)
            print(ArrivalTime)
            print(FlightTime)
            print(Distance)



            testVar = request.POST.get('testVar', '')

            final_req_str2 = ""


            air_segs = """"""

            air_pasg = """"""



            no_seats = int(no_of_seats)

            for i in range(1, no_seats + 1):
                air_pasg = air_pasg + """<com:SearchPassenger Code="ADT" Key=""" + '"' + str(i) + '"' """ Age="50"/>"""


            for K, G, C, F, O, D, dt, at, ft, dis in zip(Keyy, Group, Carrier, FlightNumber, Origin, Destination,DepartureTime, ArrivalTime, FlightTime, Distance):

                    dt_obj = dateutil.parser.parse(dt).strftime('%Y-%m-%dT%H:%M:%S.%f')

                    at_obj = dateutil.parser.parse(at).strftime('%Y-%m-%dT%H:%M:%S.%f')

                    air_segs = air_segs + """<air:AirSegment Key=""" + '"' + K + '" ' + """Group=""" + '"' + G + '"' + \
                               """ Carrier= """ + '"' + C + '"' + """ FlightNumber= """ + '"' + F + '"' + """ Origin=""" + '"' + O + '"' + \
                               """ Destination= """ + '"' + D + '"' + """ DepartureTime= """ + '"' + dt_obj + '"' + \
                               """ ArrivalTime= """ + '"' + at_obj + '"' + """ FlightTime= """ + '"' + ft + '"' + \
                               """ Distance= """ + '"' + dis + '"' + """ ETicketability= """ + '"' + """Yes""" + '"' + \
                               """ Equipment=""" + '"' + """ATR""" + '"' + """ ChangeOfPlane= """ + '"' + """false""" + '"' + \
                               """ OptionalServicesIndicator= """ + '"' + """false""" + '"' + """ AvailabilityDisplayType= """ + '"' + """Fare Shop/Optimal Shop""" + '"' + \
                               """ ProviderCode=""" + '"' + """1G""" + '"' + """>""" + """ </air:AirSegment>"""
                    print("printing air segment.********")
                    print(air_segs)

            try:

                payload = """
                                   <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/">
  <soapenv:Header />
  <soapenv:Body>
    <air:AirPriceReq xmlns:com="http://www.travelport.com/schema/common_v50_0" xmlns:air="http://www.travelport.com/schema/air_v50_0" TargetBranch="P7009927" CheckOBFees="All" AuthorizedBy="user" TraceId="3BCE9B92">
      <com:BillingPointOfSaleInfo OriginApplication="UAPI" />
      <air:AirItinerary>

    """ + air_segs + """</air:AirItinerary>"""  +"""<air:AirPricingModifiers FaresIndicator="AllFares">

   </air:AirPricingModifiers>""" +   air_pasg    + """<air:AirPricingCommand/>   </air:AirPriceReq>  </soapenv:Body></soapenv:Envelope>"""

                header = {
                    "Content-Type": "text/xml:charset=utf-8",
                    "Accept": "gzip,deflate",
                    "Connection": "Keep-Alive",
                    "Authorization": "Basic %s" % CREDENTIALS_enc64,
                    "Content-Length": str(len(payload))
                }

                url = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/AirService"
                print(payload)
                response = requests.post(url, data=payload, headers=header)
                fligh_response = response.text
                print(fligh_response)

                handler = FlightDetailsHandler()

                xml.sax.parseString(fligh_response, handler)

                var1 = handler.getFinalData()

                var2 = json.dumps(var1)

                var3 = json.loads(var2)

                print("printing dataaaa")

                print(type(var1))

                print(type(var2))

                print(type(var3))

                if "error_flag" in var3:

                    print("printing error flag")
                    print(var3["error_flag"])
                    msg = "Error in flight booking"

                    messages.error(request, str(msg))

                    if "faultcode" in var3["error_flag"] and var3["error_flag"] == "400":
                        msg = "Request data not found - Request ignored"
                        messages.error(request, str(msg))

                    if "faultstring" in var3:
                        msg2 = var3["faultstring"]
                        messages.error(request, str(msg2))

                else:

                    #################################################

                    result1 = fligh_response.find('<air:AirItinerary')

                    result2 = fligh_response.find('</air:AirItinerary>')

                    print('Printing sub treeeeeeeeeeeee')

                    # print(fligh_response[result1:result2])

                    new_req_tr = fligh_response[result1:result2]

                    final_req_str = new_req_tr.replace("<air:AirItinerary>", "")

                    final_req_str2 = final_req_str.replace("</air:AirItinerary>", "")

                    # AirPricingSolutionStr = new_req_tr.replace("air:","ns2:")

                    AirPricingSolutionStr = ""

                    new_air_pricing_resp = ""

                    r1 = fligh_response.find('<air:AirPricingSolution')

                    #r2 = fligh_response.find('</air:AirPriceResult>')

                    r2 = fligh_response.find('</air:AirPricingSolution>')

                    air_pricing_resp = fligh_response[r1:r2]

                    #aps = '</air:AirPricingSolution>'

                    myTuple = ( air_pricing_resp , "</air:AirPricingSolution>" )

                    air_pricing_resp = ''.join(myTuple)

                    print("************** AirPricingSolutionStr String **********")

                    print(air_pricing_resp)

                    print('*************** Air Segment String ****************')
                    print(final_req_str2)


                    new_air_pricing_resp = air_pricing_resp


                    if (new_air_pricing_resp.find('<air:AirPricingSolution') != -1):

                        print("Contains given substring ")

                        result1 = new_air_pricing_resp.find('<air:AirPricingSolution')

                        print(result1)

                        result2 = new_air_pricing_resp.find('>', result1)

                        print(result2)

                        new_air_pricing_resp23 = new_air_pricing_resp[result1:result2 + 1]

                        print(new_air_pricing_resp23)

                        new_air_pricing_resp2 = new_air_pricing_resp[result1:result2 + 1]

                        new_air_pricing_resp3 = new_air_pricing_resp2.replace('>',
                                                                              ' xmlns:air="http://www.travelport.com/schema/air_v50_0" >')

                        print(new_air_pricing_resp3)

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp23,
                                                                            new_air_pricing_resp3)

                    else:
                        print("Doesn't contains given substring")
                        # new_air_pricing_resp3 = new_air_pricing_resp

                    count = 0
                    start = 0

                    while (new_air_pricing_resp.find('<common_v50_0:Endorsement', start) != -1):
                        result1 = new_air_pricing_resp.find('<common_v50_0:Endorsement', start)

                        result2 = new_air_pricing_resp.find('/>', result1)

                        start = result2

                        count = count + 1

                        new_air_pricing_resp44 = new_air_pricing_resp[result1:result2 + 2]

                        new_air_pricing_resp4 = new_air_pricing_resp[result1:result2 + 2]

                        if (new_air_pricing_resp4.find('xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0"') == -1 ):

                            new_air_pricing_resp5 = new_air_pricing_resp4.replace('/>',' xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0" />')
                        else:
                            new_air_pricing_resp5 = new_air_pricing_resp4

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp44,new_air_pricing_resp5)

                    count3 = 0
                    start3 = 0


                    while (new_air_pricing_resp.find('<common_v50_0:HostToken',start3) != -1):

                        result1 = new_air_pricing_resp.find('<common_v50_0:HostToken',start3)

                        result2 = new_air_pricing_resp.find('>', result1)

                        start3 = result2

                        count3 = count3 + 1

                        new_air_pricing_resp11 = new_air_pricing_resp[result1:result2 + 1]

                        new_air_pricing_resp22 = new_air_pricing_resp[result1:result2 + 1]

                        if (new_air_pricing_resp22.find('xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0"') == -1):

                            new_air_pricing_resp33 = new_air_pricing_resp22.replace('>',' xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0" >')
                        else:
                            new_air_pricing_resp33 = new_air_pricing_resp22

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp11,new_air_pricing_resp33)



                    count3 = 0
                    start3 = 0

                    while (new_air_pricing_resp.find('<common_v50_0:ServiceData', start3) != -1):

                        result1 = new_air_pricing_resp.find('<common_v50_0:ServiceData', start3)

                        result2 = new_air_pricing_resp.find('>', result1)

                        start3 = result2

                        count3 = count3 + 1

                        new_air_pricing_resp11 = new_air_pricing_resp[result1:result2 + 1]

                        new_air_pricing_resp22 = new_air_pricing_resp[result1:result2 + 1]

                        if (new_air_pricing_resp22.find(
                                'xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0"') == -1):

                            new_air_pricing_resp33 = new_air_pricing_resp22.replace('/>',
                                                                                    ' xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0" />')
                        else:
                            new_air_pricing_resp33 = new_air_pricing_resp22

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp11,new_air_pricing_resp33)



                    count3 = 0
                    start3 = 0

                    while (new_air_pricing_resp.find('<common_v50_0:ServiceInfo', start3) != -1):

                        result1 = new_air_pricing_resp.find('<common_v50_0:ServiceInfo', start3)

                        result2 = new_air_pricing_resp.find('>', result1)

                        start3 = result2

                        count3 = count3 + 1

                        new_air_pricing_resp11 = new_air_pricing_resp[result1:result2 + 1]

                        new_air_pricing_resp22 = new_air_pricing_resp[result1:result2 + 1]

                        if (new_air_pricing_resp22.find(
                                'xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0"') == -1):

                            new_air_pricing_resp33 = new_air_pricing_resp22.replace('>',
                                                                                    ' xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0" >')
                        else:
                            new_air_pricing_resp33 = new_air_pricing_resp22

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp11,new_air_pricing_resp33)



                    count3 = 0
                    start3 = 0

                    while (new_air_pricing_resp.find('<common_v50_0:Remarks', start3) != -1):

                        result1 = new_air_pricing_resp.find('<common_v50_0:Remarks', start3)

                        result2 = new_air_pricing_resp.find('>', result1)

                        start3 = result2

                        count3 = count3 + 1

                        new_air_pricing_resp11 = new_air_pricing_resp[result1:result2 + 1]

                        new_air_pricing_resp22 = new_air_pricing_resp[result1:result2 + 1]

                        if (new_air_pricing_resp22.find(
                                'xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0"') == -1):

                            new_air_pricing_resp33 = new_air_pricing_resp22.replace('>',
                                                                                    ' xmlns:common_v50_0="http://www.travelport.com/schema/common_v50_0" >')
                        else:
                            new_air_pricing_resp33 = new_air_pricing_resp22

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp11,new_air_pricing_resp33)



                    start11 = 0
                    count11 = 0

                    while (new_air_pricing_resp.find('<air:AirSegmentRef', start11) != -1):

                        result1 = new_air_pricing_resp.find('<air:AirSegmentRef', start11)

                        result2 = new_air_pricing_resp.find('/>', result1)

                        start11 = result2

                        count11 = count11 + 1

                        new_air_pricing_resp6 = new_air_pricing_resp[result1:result2 + 2]

                        if count11 == 1:

                            new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp6, final_req_str2)

                        else:

                            new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp6, '')




                    start11 = 0
                    count11 = 0

                    while (new_air_pricing_resp.find('<air:AirSegmentRef', start11) != -1):

                        result1 = new_air_pricing_resp.find('<air:AirSegmentRef', start11)

                        result2 = new_air_pricing_resp.find('/>', result1)

                        start11 = result2

                        count11 = count11 + 1

                        new_air_pricing_resp6 = new_air_pricing_resp[result1:result2 + 2]

                        new_air_pricing_resp = new_air_pricing_resp.replace(new_air_pricing_resp6, '')






                request = get_request()
                login_type = request.session['admin_login_type']
                access_token = request.session['admin_access_token']
                headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

                url_emp = settings.API_BASE_URL + "employees"
                payload = {'corporate_id': 1 , 'spoc_id': request.user.id}
                r = requests.post(url_emp, data=payload, headers=headers)
                company_emp = json.loads(r.text)
                employees = company_emp['Employees']

                url_enty = settings.API_BASE_URL + "billing_entities"
                entys = getDataFromAPI(login_type, access_token, url_enty, payload)
                entities = entys['Entitys']

                url_ass_code = settings.API_BASE_URL + "get_assessment_code"
                ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
                ass_code = ass_code['AssCodes']

                url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
                cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
                cities_ass = cities_ass['AssCity']

                url_access = settings.API_BASE_URL + "view_company"
                data = getDataFromAPI(login_type, access_token, url_access, payload)
                access = data['Corporates']

                url_spoc = settings.API_BASE_URL + "spocs"
                spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
                spocs = spoc['Spocs']

                url_access = settings.API_BASE_URL + "get_airports"
                data = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data['Airports']


                ##############################################


                return render(request, 'Company/Admin/add_flight_booking_conformation.html', { "response": var3 , "UID2": 1 , 'employees':employees,'entities':entities,
                            'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs,'airports':airports , "booking_datas" : booking_datas , "AirSegString" : final_req_str2 , 'AirPricingSolution': new_air_pricing_resp } )

            except Exception as e:
                messages.error(request, e)
                print(traceback.print_exc())
                return redirect('travelport_flight_book')
        else:

            if 'admin_login_type' in request.session:
                login_type = request.session['admin_login_type']
                access_token = request.session['admin_access_token']
                payload = {'corporate_id': id}

                ###########################################

                request = get_request()
                login_type = request.session['admin_login_type']
                access_token = request.session['admin_access_token']
                headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

                url_emp = settings.API_BASE_URL + "employees"
                payload = {'corporate_id': 1, 'spoc_id': request.user.id}
                r = requests.post(url_emp, data=payload, headers=headers)
                company_emp = json.loads(r.text)
                employees = company_emp['Employees']

                url_enty = settings.API_BASE_URL + "billing_entities"
                entys = getDataFromAPI(login_type, access_token, url_enty, payload)
                entities = entys['Entitys']

                url_ass_code = settings.API_BASE_URL + "get_assessment_code"
                ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
                ass_code = ass_code['AssCodes']

                url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
                cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
                cities_ass = cities_ass['AssCity']

                url_access = settings.API_BASE_URL + "view_company"
                data = getDataFromAPI(login_type, access_token, url_access, payload)
                access = data['Corporates']

                url_spoc = settings.API_BASE_URL + "spocs"
                spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
                spocs = spoc['Spocs']

                url_access = settings.API_BASE_URL + "get_airports"
                data = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data['Airports']

                ##############################################


            return render(request, 'Company/Admin/add_flight_booking_conformation.html',{'employees':employees,'entities':entities,
                            'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs,'airports':airports } )

    else:
        return HttpResponseRedirect("/login")


def travelport_flight_search(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':

            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            payload = {'corporate_id': id}

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            user_id = request.POST.get('admin_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            trip_type = request.POST.get('trip_type', '')

            return_date = request.POST.get('return_date', '')

            fl_class = request.POST.get('fl_class', '')
            no_of_seats = request.POST.get('no_of_seats', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            departure_date = request.POST.get('departure_date', '')

            dep_data = dateutil.parser.parse(departure_date).strftime('%Y-%m-%d')

            #dep_data2 = dateutil.parser.parse(departure_date).strftime('%m/%d/%y')

            oneway = 0
            roundtrip = 0
            return_date_data = ''

            booking_datas = { "from_city" : from_city , "to_city" : to_city , "departure_datetime" : departure_date ,"return_date" : return_date , "fl_class" : fl_class , "no_of_seats" : no_of_seats , "trip_type" : trip_type }

            print("Printing Booking Dataaaa **************")
            print(booking_datas)

            if trip_type == '2' or trip_type == 2:
                return_data = dateutil.parser.parse(return_date).strftime('%Y-%m-%d')
                roundtrip = 1
                return_date_data = """
                            <air:SearchAirLeg>
                                <air:SearchOrigin>
                                  <com:CityOrAirport Code=""" + '"' + to_city + '"' + """ />
                                </air:SearchOrigin>
                                <air:SearchDestination>
                                  <com:CityOrAirport Code=""" + '"' + from_city + '"' + """ />
                                </air:SearchDestination>
                                <air:SearchDepTime PreferredTime=""" + '"' + return_data + '"' + """ />
                                <air:AirLegModifiers>
                                  <air:PermittedCabins>
                                    <CabinClass Type=""" + '"' + fl_class + '"' + """ xmlns="http://www.travelport.com/schema/common_v50_0" />
                                  </air:PermittedCabins>
                                </air:AirLegModifiers>
                              </air:SearchAirLeg>
                           """
            else:
                oneway = 1
            no_of_pass_str = '<com:SearchPassenger Code="ADT" />'
            n = int(no_of_seats)
            print("printtt n ********")
            print(n)
            pass_no_str = []
            no_of_pass_str = ""
            for i in range(n):
                pass_no_str.append('<com:SearchPassenger Code="ADT" />')

            print("Passanger array *********************")
            print(pass_no_str)



            for passg in pass_no_str :
                no_of_pass_str += passg

            booking_data = {'user_id': user_id, 'user_type': login_type, 'corporate_id': corporate_id,
                            'spoc_id': spoc_id, 'group_id': group_id,
                            'subgroup_id': subgroup_id, 'from_city': from_city, 'to_city': to_city,
                            'departure_datetime': departure_date, 'return_date': return_date, 'trip_type': trip_type,
                            'fl_class': fl_class, 'no_of_seats': no_of_seats}

            payload = {'auth_token': "", 'session_id': access_token, 'from_city': from_city, 'to_city': to_city,
                       'departure_date': departure_date,
                       'fl_class': fl_class, 'return_date': return_date, 'trip_type': trip_type,
                       'no_of_seats': no_of_seats}
            print(payload)
            try:
                payload = """
                                    <soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/" >
                                      <soap:Body>
                                        <air:LowFareSearchReq TargetBranch=""" + '"' + RGETBRANCH + '"' + """ TraceId="3BCE9B92" SolutionResult="false" AuthorizedBy="SUSIL" xmlns:air="http://www.travelport.com/schema/air_v50_0" xmlns:com="http://www.travelport.com/schema/common_v50_0">
                                          <com:BillingPointOfSaleInfo OriginApplication="UAPI" />
                                          <air:SearchAirLeg>
                                            <air:SearchOrigin>
                                              <com:Airport Code=""" + '"' + from_city + '"' + """ />
                                            </air:SearchOrigin>
                                            <air:SearchDestination>
                                              <com:Airport Code=""" + '"' + to_city + '"' + """ />
                                            </air:SearchDestination>
                                            <air:SearchDepTime PreferredTime=""" + '"' + dep_data + '"' + """ />
                                          </air:SearchAirLeg>

                                          """ + return_date_data + """

                                           <air:AirSearchModifiers>
                                            <air:PreferredProviders>
                                              <com:Provider Code=""" + '"' + Provider + '"' + """ />
                                            </air:PreferredProviders>
                                            <air:PermittedCabins>
                                              <CabinClass Type=""" + '"' + fl_class + '"' + """ xmlns="http://www.travelport.com/schema/common_v50_0" />
                                            </air:PermittedCabins>
                                          </air:AirSearchModifiers>

                                          """ + no_of_pass_str + """
                                          
                                        </air:LowFareSearchReq>
                                      </soap:Body>
                                    </soap:Envelope>
                                  """
                header = {
                    "Content-Type": "text/xml:charset=utf-8",
                    "Accept": "gzip,deflate",
                    "Connection": "Keep-Alive",
                    "Authorization": "Basic %s" % CREDENTIALS_enc64,
                    "Content-Length": str(len(payload))
                }

                url = "https://apac.universal-api.pp.travelport.com/B2BGateway/connect/uAPI/AirService"
                print("payload *****************************")
                print(payload)
                response = requests.post(url, data=payload, headers=header)
                fligh_response = response.text
                print("api response ****************************")
                print(fligh_response)

                handler = FlightHandler2()

                xml.sax.parseString(fligh_response, handler)

                var1 = handler.getFinalData()

                var2 = json.dumps(var1)

                var3 = json.loads(var2)

                print("printing dataaaa")

                if "error_flag" in var3:

                    print("printing error flag")
                    print(var3["error_flag"])
                    msg = "Error In Flight Search"

                    messages.error(request, str(msg))

                    if "faultcode" in var3["error_flag"] and var3["error_flag"] == "400" :
                        msg = "Request data not found - Request ignored"
                        messages.error(request, str(msg))

                    if "faultstring" in var3:
                        msg2 = var3["faultstring"]
                        messages.error(request, str(msg2))

                    return render(request, 'Company/Admin/add_flight_booking_serarch_travelport_api.html',
                                  {"response": {}, 'airports': airports, "oneway": oneway, 'roundtrip': roundtrip,
                                   "booking_datas": booking_datas})

                else:

                    return render(request, 'Company/Admin/add_flight_booking_serarch_travelport_api.html',
                                  {"response": var3, 'airports': airports, "oneway": oneway, 'roundtrip': roundtrip,
                                   "booking_datas": booking_datas})

                #print(type(var1))

                #print(type(var2))

                #print(type(var3))



            except Exception as e:
                messages.error(request, e)
                print(traceback.print_exc())
                return redirect('travelport_flight_search')
        else:

            if 'admin_login_type' in request.session:
                login_type = request.session['admin_login_type']
                access_token = request.session['admin_access_token']
                payload = {'corporate_id': id}

                url_access = settings.API_BASE_URL + "get_airports"
                data = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data['Airports']


            return render(request, 'Company/Admin/add_flight_booking_serarch_travelport_api.html',{'airports':airports} )

    else:
        return HttpResponseRedirect("/login")


############################################## GUEST HOUSE ##########################################


def guesthouse_bookings(request, id):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "company_guesthouse_booking"
        payload = {'user_id': user_id, 'request_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        # print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/guesthouse_bookings.html",{'bookings': booking, 'booking_type': id})
        else:
            return render(request, "Company/Admin/guesthouse_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def guesthouse_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "company_guesthouse_booking"
        payload = {'user_id':  request.user.id, 'request_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_guesthouse_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_guesthouse_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def guesthouse_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_guesthouse_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_guesthouse_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")

def search_guesthouse(request, id):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        payload = {'user_id': user_id, 'corporate_id': request.user.corporate_id}

        url_hotel_types = settings.API_BASE_URL + "get_guesthouse_by_corporate"
        hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)

        if hotel_types['success'] == 1:
            guesthouse = hotel_types['Guesthouse']
            return render(request, "Company/Admin/search_guesthouses.html", {'guesthouses': guesthouse})
        else:
            return render(request, "Company/Admin/search_guesthouses.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_guesthouse_booking(request, id):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_guesthouse_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_guesthouse_booking.html", {'bookings': booking})
        else:
            return render(request, "Company/Admin/view_guesthouse_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_guesthouse_booking(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "employee_reject_guesthouse_booking"

        payload = {'booking_id': booking_id, 'user_id': user_id, 'user_comment': user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Guesthouse Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Guesthouse to Reject Train Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Fails"})
    else:
        return redirect('/login')


def get_guesthouse_to_booking(request, id):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        gid = request.POST.get('guesthouse_id', '')
        from_city = request.POST.get('from_city', '')
        check_in = request.POST.get('check_in_datetime', '')
        check_out = request.POST.get('check_out_datetime', '')
        gid = request.POST.get('guesthouse_id', '')
        rid = request.POST.get('room_id', '')
        no_of_nights = request.POST.get('no_of_nights', '')

        payload = {'corporate_id': request.user.corporate_id, 'guesthouse_id': gid, 'spoc_id':request.user.id}

        url_hotel_types = settings.API_BASE_URL + "get_guesthouse_details"
        hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
        guesthouse = hotel_types['Guesthouse']

        url_city = settings.API_BASE_URL + "get_assessment_city"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        cities_ass = cities['AssCity']

        url_ass_code = settings.API_BASE_URL + "get_assessment_code"
        ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
        ass_code = ass_code['AssCodes']

        url_access = settings.API_BASE_URL + "view_company"
        data = getDataFromAPI(login_type, access_token, url_access, payload)
        access = data['Corporates']

        url_emp = settings.API_BASE_URL + "spoc_employee"
        company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
        employees = company_emp['Employees']

        url_enty = settings.API_BASE_URL + "billing_entities"
        entys = getDataFromAPI(login_type, access_token, url_enty, payload)
        entities = entys['Entitys']

        url_spoc = settings.API_BASE_URL + "spocs"
        spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
        spocs = spoc['Spocs']

        if gid:
            return render(request, 'Company/Admin/add_guesthouse_booking.html',
                          {'cities_ass': cities_ass, 'guesthouses': guesthouse,'spocs':spocs,'no_of_nights':no_of_nights,
                           'assessments': ass_code, 'cities': cities, 'corp_access': access, 'from_city': from_city,
                           'check_in': check_in, 'check_out': check_out, 'gid': gid, 'rid': rid,'employees':employees,'entities':entities})
        else:
            return render(request, 'Company/Admin/add_guesthouse_booking.html', {})
    else:
        return HttpResponseRedirect("/login")


def add_guesthouse_booking(request, id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = 0
            subgroup_id = 0
            payload1 = {'spoc_id': spoc_id}
            url_get_spoc = settings.API_BASE_URL + "view_spoc"
            spoc = getDataFromAPI(login_type, access_token, url_get_spoc, payload1)
            spoc = spoc['Spoc']

            for spoc in spoc:
                group_id = spoc['group_id']
                subgroup_id = spoc['subgroup_id']

            room_id = request.POST.get('room_id')
            guesthouse_id = request.POST.get('guesthouse_id')

            city_id = request.POST.get('city_id')
            booking_date = request.POST.get('booking_datetime')
            checkin_datetime = request.POST.get('checkin_datetime', '')
            checkout_datetime = request.POST.get('checkout_datetime', '')
            no_of_nights = request.POST.get('no_of_nights', '')
            assessment_code = request.POST.get('assessment_code')
            assessment_city = request.POST.get('assessment_city')
            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')

            if billing_entity:
                pass
            else:
                billing_entity = 0

            employees = []

            no_of_seats = request.POST.get('no_of_seats')
            act_no_of_seats = int(no_of_seats) + 1
            employees = []

            for i in range(1, act_no_of_seats):
                employees.append(request.POST.get('employee_id_' + str(i), ''))

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'room_id': room_id, 'guesthouse_id': guesthouse_id,
                       'booking_datetime': booking_date, 'city_id': city_id,'no_of_seats':no_of_seats,
                       'checkin_datetime': checkin_datetime + ':00', 'checkout_datetime': checkout_datetime + ':00',
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city,
                       'billing_entity_id': billing_entity, 'employees': employees, 'no_of_nights': no_of_nights,
                       'reason_booking': reason_for_booking, 'is_sms': 1, 'is_email': 1}
            print(payload)
            url_taxi_booking = settings.API_BASE_URL + "add_guesthouse_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/guesthouse-bookings/2", {})
            else:
                messages.error(request, 'Failed To Add Guesthouse Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/guesthouse-bookings/2", {})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            gid = request.POST.get('guesthouse_id', '')

            payload = {'corporate_id': request.user.corporate_id, 'guesthouse_id': gid, 'spoc_id':request.user.id}

            url_hotel_types = settings.API_BASE_URL + "get_guesthouse_details"
            hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
            guesthouse = hotel_types['Guesthouse']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities_ass = cities['AssCity']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_emp = settings.API_BASE_URL + "spoc_employee"
            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employees']

            from_city = request.POST.get('from_city', '')
            check_in = request.POST.get('check_in_datetime', '')
            check_out = request.POST.get('check_out_datetime', '')
            gid = request.POST.get('guesthouse_id', '')
            rid = request.POST.get('room_id', '')

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if rid:
                return render(request, 'Company/Admin/add_guesthouse_booking.html',
                              {'cities_ass': cities_ass, 'guesthouses': guesthouse,'spocs':spocs,
                               'assessments': ass_code, 'cities': cities, 'corp_access': access, 'from_city': from_city,
                               'check_in': check_in, 'check_out': check_out, 'gid': gid, 'rid': rid,'employees':employees,'entities':entities})
            else:
                return render(request, 'Company/Admin/add_guesthouse_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


def view_guesthouse_details(request, id):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        payload = {'corporate_id': request.user.corporate_id, 'guesthouse_id': id}

        url_hotel_types = settings.API_BASE_URL + "get_guesthouse_details"
        hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
        guesthouse = hotel_types['Guesthouse']

        if id:
            return render(request, 'Company/Admin/view_guesthouse_details.html', {'guesthouses': guesthouse})
        else:
            return render(request, 'Company/Admin/view_guesthouse_details.html', {})
    else:
        return HttpResponseRedirect("/login")


def download_guesthouse_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_guesthouse_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name, 'user_id': user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Guesthouse-Bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Guesthouse Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Guesthouse Code',
        'Guesthouse Name',
        'Room Number',
        'Room Name',
        'Room Code',
        'Preferred Hotel',

        'TaxiVaxi Status',
        'Employees Name',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"
        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            'Billing Entity',
            'Travel request Code',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_city_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            bk['guest_house_code'],
            bk['guest_house_name'],
            bk['room_number'],
            bk['room_name'],
            bk['room_code'],

            bk['preferred_hotel'],

            bk['status_cotrav'],
            passanger_list,

            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


